import json
from collections.abc import Mapping, Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models import (
    AgricultureRelatedResult,
    FiveArticlesResult,
    InclusiveFinanceResult,
    NationalEconomyClassificationCase,
    NationalEconomyClassificationResult,
)
from app.services.national_economy_case_ingestion import FIELD_LABELS
from app.services.national_economy_classification_workflow import (
    get_current_completed_result,
)
from app.services.national_economy_result_presentation import (
    format_industry_display_code,
)
from app.services.inclusive_finance_determination import BORROWER_TYPE_LABELS
from app.services.scenario_registry import (
    INCLUSIVE_FINANCE_FIELD_SCHEMA,
    SCENARIO_REGISTRY,
    ScenarioRegistration,
    TECHNOLOGY_FINANCE_SCENARIO,
)


CASE_INPUT_SHEET = "案例输入"
CURRENT_RESULT_SHEET = "当前结论"
RESULT_HISTORY_SHEET = "判定历史"
TECHNOLOGY_FINANCE_RESULT_SHEET = "科技金融判定"
INCLUSIVE_FINANCE_RESULT_SHEET = "普惠金融判定"
AGRICULTURE_RELATED_RESULT_SHEET = "涉农判定"
TECHNOLOGY_FINANCE_CONSISTENCY_LABEL = (
    "贷款对应的五篇大文章类别与企业类别是否一致"
)
_IP_INTENSIVE_INDUSTRY_SUBJECTS = frozenset(
    {"知识产权（专利）密集型产业", "知识产权(专利)密集型产业"}
)

_TECHNOLOGY_FINANCE_HEADERS = (
    "Stage B版本",
    "科技金融状态",
    "状态说明",
    "Stage A结果ID",
    "贷款投向国民经济行业代码",
    "贷款投向国民经济行业名称",
    "企业国民经济行业代码",
    "企业国民经济行业名称",
    "主题",
    "第一层",
    "第二层",
    "第三层",
    "第四层",
    "映射代码",
    "映射名称",
    "映射源行",
    "匹配依据",
    "业务证据摘要",
    "知识产权条件",
    TECHNOLOGY_FINANCE_CONSISTENCY_LABEL,
    "一致性依据",
)

_RESULT_STATUS_LABELS = {
    "completed": "判定完成",
    "not_applicable": "不属于科技金融",
    "needs_review": "待人工复核",
    "classification_failed": "判定失败",
}

_CONSISTENCY_STATUS_LABELS = {
    "consistent": "一致",
    "inconsistent": "不一致",
    "needs_review": "待人工复核",
    "not_applicable": "不适用",
}

_AGRICULTURE_STATUS_LABELS = {
    "completed": "判定完成",
    "not_applicable": "不属于涉农",
    "needs_review": "待人工复核",
    "classification_failed": "判定失败",
}

_AGRICULTURE_RESULT_HEADERS = (
    "版本号",
    "状态",
    "状态说明",
    "命中类别",
    "是否涉农",
    "匹配依据",
    "各类别判定方式",
    "创建时间",
)


def export_case_workbook(
    case: NationalEconomyClassificationCase,
    *,
    five_articles_results: Sequence[FiveArticlesResult] = (),
    profile: ScenarioRegistration | None = None,
    inclusive_finance_results: Sequence[InclusiveFinanceResult] = (),
    agriculture_related_results: Sequence[AgricultureRelatedResult] = (),
) -> bytes:
    workbook = Workbook()
    input_sheet = workbook.active
    input_sheet.title = CASE_INPUT_SHEET
    _write_case_input(input_sheet, case)

    current_sheet = workbook.create_sheet(CURRENT_RESULT_SHEET)
    _write_current_result(current_sheet, case)

    history_sheet = workbook.create_sheet(RESULT_HISTORY_SHEET)
    _write_result_history(history_sheet, case)

    resolved_profile = profile or SCENARIO_REGISTRY.get(case.scenario)
    if resolved_profile is not None and resolved_profile.workflow is not None:
        if resolved_profile.id != case.scenario:
            raise ValueError("案例与导出场景 profile 不一致")
        result_sheet = workbook.create_sheet(resolved_profile.export_sheet_name)
        if resolved_profile.id == "inclusive_finance":
            _write_inclusive_finance_result(result_sheet, inclusive_finance_results)
        elif resolved_profile.id == "agriculture_related":
            _write_agriculture_related_result(result_sheet, agriculture_related_results)
        elif resolved_profile.id == TECHNOLOGY_FINANCE_SCENARIO:
            _write_technology_finance_result(result_sheet, five_articles_results)
        else:
            _write_five_articles_result(
                result_sheet,
                five_articles_results,
                resolved_profile,
            )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_agriculture_related_result(
    sheet: Worksheet,
    results: Sequence[AgricultureRelatedResult],
) -> None:
    sheet.append(_AGRICULTURE_RESULT_HEADERS)
    if not results:
        sheet.append(
            (
                "",
                "尚未判定",
                "尚无涉农判定结果。",
                "",
                "未判定",
                "尚无涉农判定结果。",
                "",
                "",
            )
        )
        return

    for result in sorted(results, key=lambda item: (item.version, item.id or 0)):
        status = _AGRICULTURE_STATUS_LABELS.get(result.status, result.status)
        error_detail = result.error_detail or "未提供原因。"
        if result.status == "needs_review":
            status_detail = f"涉农判定需人工复核：{result.basis or '未提供原因。'}"
        elif result.status == "classification_failed":
            status_detail = f"涉农判定失败：{error_detail}"
        elif result.status == "not_applicable":
            status_detail = "四类判定均未命中，不属于涉农。"
        else:
            status_detail = "涉农判定完成。"

        categories = [
            item
            for item in (result.matched_categories or [])
            if isinstance(item, dict)
        ]
        matched_names = [
            str(item.get("category_name") or f"类别{item.get('category', '')}")
            for item in categories
            if item.get("result") == "matched"
        ]
        methods = [
            f"类别{item.get('category', '')}（{item.get('category_name', '未命名')}）："
            f"{_agriculture_method_label(item.get('method'))}"
            for item in categories
        ]
        basis = result.basis or (
            error_detail
            if result.status in {"needs_review", "classification_failed"}
            else "无命中类别。"
        )
        is_related = (
            "是"
            if result.is_agriculture_related is True
            else "否"
            if result.is_agriculture_related is False
            else "未判定"
        )
        created_at = result.created_at.isoformat() if result.created_at else ""
        sheet.append(
            (
                result.version,
                status,
                status_detail,
                "、".join(matched_names) or "无",
                is_related,
                basis,
                "；".join(methods) or "尚无类别判定明细。",
                created_at,
            )
        )


def _agriculture_method_label(method: object) -> str:
    return {
        "rule": "规则",
        "stage_a": "Stage A",
        "ai": "AI",
    }.get(str(method), str(method) if method else "未记录")

def _write_inclusive_finance_result(sheet: Worksheet, results: Sequence[InclusiveFinanceResult]) -> None:
    headers = (
        "Stage B版本", "普惠状态", "借款主体", "主体条件", "计算划型",
        "填报划型", "划型一致性", "是否经营性", "结构化授信额度(万元)",
        "审批意见批复额度(万元)", "最终采用额度(万元)", "额度来源及一致性",
        "是否属于普惠", "普惠子类别", "注册地址辅助", "最终判定依据",
        "业务证据摘要", "异常",
    )
    sheet.append(headers)
    if not results:
        empty_row = [""] * len(headers)
        empty_row[1] = "尚未判定"
        empty_row[headers.index("最终判定依据")] = "尚无普惠金融判定结果。"
        sheet.append(empty_row)
        return
    result = max(results, key=lambda item: (item.version, item.id or 0))
    determination = result.determination or {}
    source_labels = {
        "structured_and_approval_consistent": "结构化额度与审批意见一致",
        "structured": "仅采用结构化授信额度",
        "approval_opinion": "仅采用授信审批意见批复额度",
        "approval_opinion_multiple": "审批意见存在多个批复额度",
        "conflict": "结构化额度与审批意见冲突",
        "missing": "两处均无明确额度",
    }
    field_labels = {field.key: field.label for field in INCLUSIVE_FINANCE_FIELD_SCHEMA}
    evidence = "\n".join(
        f"{ref.get('field_label') or field_labels.get(str(ref.get('field_key') or ref.get('field')), ref.get('field_key', ref.get('field', '证据')))}：{ref.get('raw_value', '')}"
        for ref in result.evidence_refs
        if isinstance(ref, dict)
    )
    anomalies = "\n".join(str(item.get("message", item)) for item in result.anomalies if isinstance(item, dict))
    approval_amounts = determination.get("approval_credit_amounts_wan") or []
    credit_source = str(determination.get("credit_amount_source") or "missing")
    sheet.append(
        (
            result.version,
            _RESULT_STATUS_LABELS.get(result.status, result.status),
            BORROWER_TYPE_LABELS.get(result.borrower_type or "", result.borrower_type),
            determination.get("borrower_type_basis"),
            result.computed_size,
            result.filled_size,
            "一致" if result.size_consistent else ("不一致" if result.size_consistent is False else "未判定"),
            "是" if result.is_operating_loan else ("否" if result.is_operating_loan is False else "未判定"),
            determination.get("structured_credit_amount_wan"),
            "、".join(f"{float(value):g}" for value in approval_amounts),
            result.credit_amount_wan,
            source_labels.get(credit_source, credit_source),
            "是" if result.qualifies else ("否" if result.qualifies is False else "未判定"),
            result.inclusive_category,
            determination.get("farmer_registration_address_support"),
            result.basis,
            evidence,
            anomalies,
        )
    )


def _write_technology_finance_result(
    sheet: Worksheet,
    results: Sequence[FiveArticlesResult],
) -> None:
    sheet.append(_TECHNOLOGY_FINANCE_HEADERS)
    if not results:
        sheet.append(
            (
                "",
                "尚未判定",
                "尚无科技金融判定结果。",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "不适用",
                "尚无科技金融判定结果，一致性不适用。",
            )
        )
        return

    current = max(results, key=lambda result: (result.version, result.id or 0))
    status_label = _RESULT_STATUS_LABELS.get(current.status, current.status)
    status_detail = _technology_finance_status_detail(current)
    consistency_label = _consistency_status_label(current)
    consistency_basis = _consistency_basis(current)
    common_values = (
        current.version,
        status_label,
        status_detail,
        current.stage_a_result_id,
        _cell_value(current.loan_neic_code),
        _cell_value(current.loan_neic_name),
        _cell_value(current.enterprise_neic_code),
        _cell_value(current.enterprise_neic_name),
    )

    if current.status != "completed" or not current.labels:
        sheet.append(
            (
                *common_values,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                consistency_label,
                consistency_basis,
            )
        )
        return

    for label in current.labels:
        taxonomy = _label_taxonomy(label)
        sheet.append(
            (
                *common_values,
                *taxonomy,
                _cell_value(label.get("NEIC_Code", label.get("code"))),
                _cell_value(label.get("NEIC_Name", label.get("name"))),
                _cell_value(label.get("source_row")),
                _cell_value(label.get("matching_basis")),
                _business_evidence_summary(label.get("evidence_refs")),
                _ip_intensive_industry_condition(label),
                consistency_label,
                consistency_basis,
            )
        )


def _ip_intensive_industry_condition(label: Mapping[str, object]) -> str:
    if label.get("subject") not in _IP_INTENSIVE_INDUSTRY_SUBJECTS:
        return ""
    status = label.get("ip_intensive_industry_status")
    if status == "satisfied":
        return "满足"
    if status == "unsatisfied":
        basis = _cell_value(label.get("ip_intensive_industry_basis"))
        return f"不满足：{basis}" if basis else "不满足"
    return ""


def _write_five_articles_result(
    sheet: Worksheet,
    results: Sequence[FiveArticlesResult],
    profile: ScenarioRegistration,
) -> None:
    sheet.append(_five_articles_headers(profile))
    if not results:
        sheet.append(
            (
                "",
                "尚未判定",
                f"尚无{profile.name}判定结果。",
                *("" for _ in range(15)),
                "不适用",
                f"尚无{profile.name}判定结果，一致性不适用。",
                *_green_export_values(None, profile),
                *_digital_export_values(None, profile),
                *_pension_export_values(None, profile),
            )
        )
        return

    current = max(results, key=lambda result: (result.version, result.id or 0))
    status_label = _five_articles_status_label(current, profile)
    status_detail = _five_articles_status_detail(current, profile)
    consistency_label = _consistency_status_label(current)
    consistency_basis = _five_articles_consistency_basis(current, profile)
    common_values = (
        current.version,
        status_label,
        status_detail,
        current.stage_a_result_id,
        _cell_value(current.loan_neic_code),
        _cell_value(current.loan_neic_name),
        _cell_value(current.enterprise_neic_code),
        _cell_value(current.enterprise_neic_name),
    )

    if (
        current.status != "completed"
        and not (profile.id == "green_finance" and current.labels)
    ) or not current.labels:
        sheet.append(
            (
                *common_values,
                *("" for _ in range(10)),
                consistency_label,
                consistency_basis,
                *_green_export_values(current, profile),
                *_digital_export_values(current, profile),
                *_pension_export_values(current, profile),
            )
        )
        return

    for label in current.labels:
        sheet.append(
            (
                *common_values,
                *_label_taxonomy(label),
                _cell_value(label.get("NEIC_Code", label.get("code"))),
                _cell_value(label.get("NEIC_Name", label.get("name"))),
                _cell_value(label.get("source_row")),
                _cell_value(label.get("matching_basis")),
                _business_evidence_summary(label.get("evidence_refs")),
                consistency_label,
                consistency_basis,
                *_green_export_values(current, profile, label),
                *_digital_export_values(current, profile, label),
                *_pension_export_values(current, profile),
            )
        )


def _five_articles_headers(profile: ScenarioRegistration) -> tuple[str, ...]:
    return (
        "Stage B版本",
        f"{profile.name}状态",
        "状态说明",
        "Stage A结果ID",
        "贷款投向国民经济行业代码",
        "贷款投向国民经济行业名称",
        "企业国民经济行业代码",
        "企业国民经济行业名称",
        "主题",
        "第一层",
        "第二层",
        "第三层",
        "第四层",
        "映射代码",
        "映射名称",
        "映射源行",
        "匹配依据",
        "业务证据摘要",
        "贷款对应的五篇大文章类别与企业类别是否一致",
        "一致性依据",
        *(
            (
                "绿色目录标签",
                "条件匹配方式",
                "绿色决策策略版本",
                "环保与绿色资质认证原文",
                "节能减排污染治理原文",
                "碳排放与环境效益原文",
                "重大环保违法失信原文",
                "重大环保违法失信状态",
                "辅助证据预警",
            )
            if profile.id == "green_finance"
            else ()
        ),
        *(
            (
                "数字类别",
                "数字决策策略版本",
                "行业定位原文",
                "数字核心竞争力原文",
                "研发知识产权原文",
                "辅助证据预警",
            )
            if profile.id == "digital_finance"
            else ()
        ),
        *(
            (
                "养老矩阵分支",
                "贷款养老投向占比原始值",
                "贷款养老投向占比规范化",
                "主营业务及营收占比原始值",
                "主营业务及营收占比规范化",
                "主体辅助依据",
                "养老资质预警",
            )
            if profile.id == "pension_finance"
            else ()
        ),
    )


def _green_export_values(
    result: FiveArticlesResult | None,
    profile: ScenarioRegistration,
    label: Mapping[str, object] | None = None,
) -> tuple[object, ...]:
    if profile.id != "green_finance":
        return ()
    if result is None:
        return ("", "", "", "", "", "", "", "", "")

    direction = next(
        (
            ref
            for ref in result.consistency_evidence_refs
            if ref.get("type") == "green_direction"
        ),
        {},
    )
    auxiliary = {
        str(ref.get("field_key")): ref
        for ref in result.consistency_evidence_refs
        if ref.get("type") == "green_auxiliary"
    }
    violation = next(
        (
            ref
            for ref in result.consistency_evidence_refs
            if ref.get("type") == "green_violation"
        ),
        {},
    )
    warnings = [
        str(ref["warning"])
        for ref in (*auxiliary.values(), violation)
        if ref.get("warning")
    ]
    taxonomy_path = (label or {}).get("taxonomy_path") or direction.get(
        "taxonomy_path"
    )
    path_values = taxonomy_path if isinstance(taxonomy_path, (list, tuple)) else ()
    directory_label = " / ".join(
        str(value)
        for value in (
            (label or {}).get("subject") or direction.get("subject"),
            *path_values,
        )
        if value
    )
    match_method = (label or {}).get("match_method") or direction.get("match_method")
    return (
        directory_label,
        "条件回退命中" if match_method == "condition_fallback" else "行业编码命中",
        _cell_value(result.decision_policy_version),
        _cell_value(auxiliary.get("green_certifications", {}).get("excerpt")),
        _cell_value(
            auxiliary.get("energy_saving_pollution_control", {}).get("excerpt")
        ),
        _cell_value(
            auxiliary.get("carbon_environmental_benefits", {}).get("excerpt")
        ),
        _cell_value(violation.get("raw_value")),
        _cell_value(violation.get("violation_status")),
        "；".join(warnings),
    )


def _digital_export_values(
    result: FiveArticlesResult | None,
    profile: ScenarioRegistration,
    label: Mapping[str, object] | None = None,
) -> tuple[object, ...]:
    if profile.id != "digital_finance":
        return ()
    if result is None:
        return ("", "", "", "", "", "")

    auxiliary_refs = {
        str(ref.get("evidence_role")): ref
        for ref in result.consistency_evidence_refs
        if ref.get("type") == "digital_auxiliary"
    }
    direction_ref = next(
        (
            ref
            for ref in result.consistency_evidence_refs
            if ref.get("type") == "digital_direction"
        ),
        {},
    )
    warnings = [
        str(ref["warning"])
        for ref in auxiliary_refs.values()
        if ref.get("warning")
    ]
    return (
        _cell_value(
            (label or {}).get("digital_category")
            or direction_ref.get("digital_category")
        ),
        _cell_value(result.decision_policy_version),
        _cell_value(auxiliary_refs.get("industry_positioning", {}).get("excerpt")),
        _cell_value(auxiliary_refs.get("core_competitiveness", {}).get("excerpt")),
        _cell_value(auxiliary_refs.get("rd_ip", {}).get("excerpt")),
        "；".join(warnings),
    )


def _pension_export_values(
    result: FiveArticlesResult | None,
    profile: ScenarioRegistration,
) -> tuple[object, ...]:
    if profile.id != "pension_finance":
        return ()
    if result is None:
        return ("", "", "", "", "", "", "")
    matrix_refs = {
        ref.get("field_key"): ref
        for ref in result.consistency_evidence_refs
        if ref.get("type") == "pension_matrix"
    }
    loan_share = matrix_refs.get("pension_loan_direction_share", {})
    revenue_share = matrix_refs.get("main_business_revenue_share", {})
    qualification = next(
        (
            ref
            for ref in result.consistency_evidence_refs
            if ref.get("type") == "pension_qualification"
        ),
        {},
    )
    branch = loan_share.get("matrix_branch") or revenue_share.get("matrix_branch")
    subject_basis = (
        "企业养老映射命中"
        if branch == "PENSION_ENTERPRISE_UNKNOWN_LOAN_SHARE"
        else "养老产业营业收入占比达到50%（含）"
        if branch == "PENSION_REVENUE_AT_LEAST_50_UNKNOWN_LOAN_SHARE"
        else "主体辅助规则未触发"
    )
    return (
        _cell_value(branch),
        _cell_value(loan_share.get("raw_value")),
        _percentage_cell(loan_share.get("normalized_percent")),
        _cell_value(revenue_share.get("raw_value")),
        _percentage_cell(revenue_share.get("normalized_percent")),
        subject_basis,
        _cell_value(qualification.get("warning")),
    )


def _percentage_cell(value: object) -> str:
    return "" if value is None else f"{value}%"


def _five_articles_status_label(
    result: FiveArticlesResult,
    profile: ScenarioRegistration,
) -> str:
    if result.status == "not_applicable":
        return f"不属于{profile.name}"
    return _RESULT_STATUS_LABELS.get(result.status, result.status)


def _five_articles_status_detail(
    result: FiveArticlesResult,
    profile: ScenarioRegistration,
) -> str:
    if result.status == "completed":
        return f"{profile.name}判定完成。"
    if result.status == "not_applicable":
        return f"贷款投向未命中已发布{profile.name}映射，不属于{profile.name}。"
    if result.status == "needs_review":
        detail = result.error_detail or result.consistency_basis
        return f"{profile.name}判定需人工复核：{detail or '映射或证据需要人工确认。'}"
    if result.status == "classification_failed":
        return f"{profile.name}判定失败：{result.error_detail or '未提供失败详情。'}"
    return result.error_detail or f"{profile.name}判定状态：{result.status}"


def _five_articles_consistency_basis(
    result: FiveArticlesResult,
    profile: ScenarioRegistration,
) -> str:
    if result.status == "completed":
        return result.consistency_basis or "未提供一致性依据。"
    if result.status == "not_applicable":
        return (
            result.consistency_basis
            or f"当前贷款投向不属于{profile.name}，一致性不适用。"
        )
    if result.status == "needs_review":
        detail = result.consistency_basis or result.error_detail
        return (
            f"当前无正式{profile.name}标签，一致性比较不适用；"
            f"{detail or '映射或证据需人工复核。'}"
        )
    if result.status == "classification_failed":
        return f"{profile.name}判定失败，未形成正式标签，一致性不适用。"
    return f"当前无正式{profile.name}标签，一致性不适用。"


def _technology_finance_status_detail(result: FiveArticlesResult) -> str:
    if result.status == "completed":
        return "科技金融判定完成。"
    if result.status == "not_applicable":
        return "贷款投向未命中已发布科技金融映射，不属于科技金融。"
    if result.status == "needs_review":
        detail = result.error_detail or result.consistency_basis
        return f"科技金融判定需人工复核：{detail or '映射或证据需要人工确认。'}"
    if result.status == "classification_failed":
        return f"科技金融判定失败：{result.error_detail or '未提供失败详情。'}"
    return result.error_detail or f"科技金融判定状态：{result.status}"


def _consistency_status_label(result: FiveArticlesResult) -> str:
    if result.status == "classification_failed":
        return "不适用"
    return _CONSISTENCY_STATUS_LABELS.get(
        result.consistency_status or "not_applicable",
        result.consistency_status or "不适用",
    )


def _consistency_basis(result: FiveArticlesResult) -> str:
    if result.status == "completed":
        return result.consistency_basis or "未提供一致性依据。"
    if result.status == "not_applicable":
        return (
            result.consistency_basis
            or "当前贷款投向不属于科技金融，一致性不适用。"
        )
    if result.status == "needs_review":
        detail = result.consistency_basis or result.error_detail
        return (
            "当前无正式科技金融标签，一致性比较不适用；"
            f"{detail or '映射或证据需人工复核。'}"
        )
    if result.status == "classification_failed":
        return "科技金融判定失败，未形成正式标签，一致性不适用。"
    return "当前无正式科技金融标签，一致性不适用。"


def _label_taxonomy(label: Mapping[str, object]) -> tuple[object, ...]:
    raw_path = label.get("taxonomy_path")
    if isinstance(raw_path, list):
        path = raw_path[:4]
        return (
            _cell_value(label.get("subject")),
            *(_cell_value(item) for item in path),
            *("" for _ in range(4 - len(path))),
        )

    raw_taxonomy = label.get("taxonomy")
    taxonomy = raw_taxonomy if isinstance(raw_taxonomy, dict) else {}
    return (
        _cell_value(taxonomy.get("subject", label.get("subject"))),
        *(_cell_value(taxonomy.get(f"tier{level}")) for level in range(1, 5)),
    )


def _business_evidence_summary(raw_refs: object) -> str:
    if not isinstance(raw_refs, list):
        return ""
    summaries = []
    for ref in raw_refs:
        if not isinstance(ref, dict) or ref.get("type") != "business":
            continue
        label = ref.get("field_label") or ref.get("field_key") or "业务证据"
        excerpt = ref.get("excerpt")
        if excerpt is not None:
            summaries.append(f"{label}：{excerpt}")
    return "\n".join(summaries)


def _write_case_input(sheet: Worksheet, case: NationalEconomyClassificationCase) -> None:
    sheet.append(("字段", "内容"))
    sheet.append(("业务场景", case.scenario))
    sheet.append(("原始文件名", case.original_filename or ""))
    sheet.append(("案例状态", case.status))
    registration = SCENARIO_REGISTRY.get(case.scenario)
    field_labels = (
        tuple((field.key, field.label) for field in registration.field_schema)
        if registration is not None
        else tuple(FIELD_LABELS.items())
    )
    for field, label in field_labels:
        sheet.append((label, _cell_value(case.input_payload.get(field))))


def _write_current_result(
    sheet: Worksheet, case: NationalEconomyClassificationCase
) -> None:
    sheet.append(
        (
            "行业代码",
            "行业名称",
            "匹配依据",
            "贷款投向代码",
            "贷款投向名称",
            "贷款投向匹配依据",
            "贷款投向是否一致",
        )
    )
    result = get_current_completed_result(case)
    if result is None:
        sheet.append(
            ("", "", f"暂无成功结论（案例状态：{case.status}）", "", "", "", "")
        )
        return
    sheet.append(_result_values(result))


def _write_result_history(
    sheet: Worksheet, case: NationalEconomyClassificationCase
) -> None:
    sheet.append(
        (
            "版本号",
            "状态",
            "行业代码",
            "行业名称",
            "匹配依据",
            "贷款投向代码",
            "贷款投向名称",
            "贷款投向匹配依据",
            "贷款投向是否一致",
            "关联异议",
        )
    )
    for result in sorted(case.result_versions, key=lambda item: item.version):
        sheet.append(
            (
                result.version,
                result.status,
                *_result_values(result),
                _objection_text(result.objection),
            )
        )


def _result_values(
    result: NationalEconomyClassificationResult,
) -> tuple[object, object, object, object, object, object, object]:
    if result.loan_industry_code is None and result.loan_matching_basis is None:
        loan_code = result.industry_code
        loan_major_code = result.industry_major_code
        loan_middle_code = result.industry_middle_code
        loan_name = result.industry_name if result.industry_code is not None else None
        loan_basis = "贷款投向未单独评估，与企业主营一致"
        loan_matches = "一致"
    elif result.loan_industry_code is None:
        loan_code = None
        loan_major_code = None
        loan_middle_code = None
        loan_name = None
        loan_basis = result.loan_matching_basis
        loan_matches = "不一致"
    else:
        loan_code = result.loan_industry_code
        loan_major_code = result.loan_industry_major_code
        loan_middle_code = result.loan_industry_middle_code
        loan_name = result.loan_industry_name
        loan_basis = result.loan_matching_basis
        loan_matches = "一致" if result.loan_matches_enterprise is True else "不一致"

    return (
        _cell_value(
            format_industry_display_code(
                result.industry_major_code,
                result.industry_code,
                result.industry_middle_code,
            )
        ),
        _cell_value(result.industry_name),
        _cell_value(result.rationale),
        _cell_value(
            format_industry_display_code(loan_major_code, loan_code, loan_middle_code)
        ),
        _cell_value(loan_name),
        _cell_value(loan_basis),
        loan_matches,
    )


def _objection_text(objection: dict[str, object] | None) -> str:
    if not objection:
        return ""
    description = objection.get("description")
    if description is not None:
        return str(description)
    return json.dumps(objection, ensure_ascii=False, sort_keys=True)


def _cell_value(value: object | None) -> object:
    return "" if value is None else value

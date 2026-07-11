from collections.abc import Callable, Sequence
from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import NationalEconomyCatalogVersion


EXPECTED_HEADERS = (
    "大类名称",
    "大类编码",
    "小类名称",
    "小类编码",
    "小类说明",
    "小类补充内容",
    "小类注释-不包括",
)


class CatalogHeaderError(ValueError):
    pass


@dataclass(frozen=True)
class CatalogSyncSource:
    path: Path
    source_hash: str
    headers: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]


@dataclass(frozen=True)
class CatalogSyncResult:
    created: bool
    resynchronized: bool
    version: NationalEconomyCatalogVersion


FullResync = Callable[[Session, NationalEconomyCatalogVersion, Sequence[tuple[Any, ...]]], None]


def read_catalog_source(path: Path) -> CatalogSyncSource:
    if not path.is_file():
        raise FileNotFoundError(f"catalog Excel not found: {path}")
    source_hash = sha256(path.read_bytes()).hexdigest()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        raw_headers = next(values, None)
        headers = tuple("" if value is None else str(value).strip() for value in raw_headers or ())
        if headers != EXPECTED_HEADERS:
            raise CatalogHeaderError(
                f"invalid catalog headers: expected {EXPECTED_HEADERS!r}, got {headers!r}"
            )
        rows = tuple(tuple(row) for row in values)
    finally:
        workbook.close()
    return CatalogSyncSource(path=path, source_hash=source_hash, headers=headers, rows=rows)


def synchronize_catalog(
    session: Session,
    source: CatalogSyncSource,
    embedding_model: str,
    embedding_dimension: int,
    full_resync: FullResync,
    *,
    force_resync: bool = False,
) -> CatalogSyncResult:
    existing_version = session.scalar(
        select(NationalEconomyCatalogVersion).where(
            NationalEconomyCatalogVersion.source_hash == source.source_hash,
            NationalEconomyCatalogVersion.embedding_model == embedding_model,
            NationalEconomyCatalogVersion.embedding_dimension == embedding_dimension,
        )
    )
    if existing_version is not None:
        if force_resync:
            full_resync(session, existing_version, source.rows)
        return CatalogSyncResult(
            created=False,
            resynchronized=force_resync,
            version=existing_version,
        )

    version = NationalEconomyCatalogVersion(
        version=sha256(
            f"{source.source_hash}:{embedding_model}:{embedding_dimension}".encode()
        ).hexdigest(),
        source_hash=source.source_hash,
        embedding_model=embedding_model,
        embedding_dimension=embedding_dimension,
    )
    session.add(version)
    session.flush()
    full_resync(session, version, source.rows)
    return CatalogSyncResult(created=True, resynchronized=True, version=version)

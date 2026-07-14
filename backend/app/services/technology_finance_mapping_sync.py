from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from hashlib import sha256
import json
import math
from pathlib import Path
import re
from typing import Any
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.models import (
    FiveArticlesMappingRow,
    FiveArticlesMappingVersion,
    NationalEconomyCatalogVersion,
    NationalEconomyIndustryChunk,
)
from app.services.scenario_registry import (
    TECHNOLOGY_FINANCE_REGISTRATION,
    ScenarioRegistration,
)


TECHNOLOGY_FINANCE_SCENARIO_ID = "technology_finance"
OPTIONAL_CATEGORY_HEADER = "属于类别"
REQUIRED_HEADERS = (
    "主题",
    "第一层名称",
    "第二层名称",
    "第三层名称",
    "第四层名称",
    "国民经济行业代码",
    "国民经济行业名称",
)
ALL_HEADERS = (OPTIONAL_CATEGORY_HEADER, *REQUIRED_HEADERS)
MAJOR_CATEGORY_CODE_PATTERN = re.compile(r"^[A-Za-z]?(\d{2})$")


class MappingHeaderError(ValueError):
    pass


@dataclass(frozen=True)
class RawMappingRow:
    source_row: int
    values: Mapping[str, Any]


@dataclass(frozen=True)
class MappingSyncSource:
    path: Path
    source_hash: str
    headers: tuple[str, ...]
    rows: tuple[RawMappingRow, ...]


@dataclass(frozen=True)
class NormalizedMappingRow:
    source_row: int
    neic_code: str
    code_level: int
    neic_name: str
    comparison_name: str
    subject: str
    tier1: str
    tier2: str | None
    tier3: str | None
    tier4: str | None

    @property
    def duplicate_key(self) -> tuple[object, ...]:
        return (
            self.neic_code,
            self.comparison_name,
            self.subject,
            self.tier1,
            self.tier2,
            self.tier3,
            self.tier4,
        )


@dataclass(frozen=True)
class CatalogFacts:
    version: NationalEconomyCatalogVersion | None
    four_digit_names: Mapping[str, frozenset[str]]
    three_digit_names: Mapping[str, frozenset[str]]
    two_digit_names: Mapping[str, frozenset[str]]
    four_digit_name_codes: Mapping[str, frozenset[str]]
    three_digit_name_codes: Mapping[str, frozenset[str]]
    two_digit_name_codes: Mapping[str, frozenset[str]]
    errors: tuple[dict[str, object], ...] = ()


@dataclass(frozen=True)
class MappingSyncResult:
    version: FiveArticlesMappingVersion
    reused: bool


def read_mapping_source(
    path: Path,
    *,
    expected_category: str | None = None,
) -> MappingSyncSource:
    if not path.is_file():
        raise FileNotFoundError(f"five-articles mapping Excel not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        raw_headers = next(values, None)
        if raw_headers is None:
            raise MappingHeaderError("five-articles mapping Excel has no header row")

        headers = tuple(_canonical_header(value) for value in raw_headers)
        positions: dict[str, int] = {}
        duplicates: list[str] = []
        for index, header in enumerate(headers):
            if not header:
                continue
            if header in positions:
                duplicates.append(header)
            else:
                positions[header] = index

        missing = [header for header in REQUIRED_HEADERS if header not in positions]
        if expected_category is not None and OPTIONAL_CATEGORY_HEADER not in positions:
            missing.append(OPTIONAL_CATEGORY_HEADER)
        relevant_duplicates = sorted(set(duplicates).intersection(ALL_HEADERS))
        if missing or relevant_duplicates:
            details = []
            if missing:
                details.append(f"missing headers: {missing!r}")
            if relevant_duplicates:
                details.append(f"duplicate headers: {relevant_duplicates!r}")
            raise MappingHeaderError("invalid mapping headers: " + "; ".join(details))

        rows = []
        for source_row, raw_row in enumerate(values, start=2):
            row_values = {
                header: raw_row[index] if index < len(raw_row) else None
                for header, index in positions.items()
                if header in ALL_HEADERS
            }
            if not any(not _is_blank(value) for value in row_values.values()):
                continue
            category = _normalize_text(row_values.get(OPTIONAL_CATEGORY_HEADER))
            if expected_category is not None and category != expected_category:
                continue
            rows.append(RawMappingRow(source_row=source_row, values=row_values))
    finally:
        workbook.close()

    return MappingSyncSource(
        path=path,
        source_hash=_source_hash(headers, rows),
        headers=headers,
        rows=tuple(rows),
    )


def _source_hash(
    headers: Sequence[str], rows: Sequence[RawMappingRow]
) -> str:
    """Hash only the selected scenario partition, not unrelated workbook rows."""
    payload = {
        "headers": list(headers),
        "rows": [
            [
                _audit_value(row.values.get(header))
                for header in ALL_HEADERS
                if header in headers
            ]
            for row in rows
        ],
    }
    return sha256(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def synchronize_technology_finance_mapping(
    session: Session,
    source: MappingSyncSource,
    settings: Settings,
    *,
    scenario_id: str = TECHNOLOGY_FINANCE_SCENARIO_ID,
) -> MappingSyncResult:
    """Compatibility wrapper for the original technology-finance sync API."""
    return _synchronize_mapping(
        session,
        source,
        settings,
        scenario_id=scenario_id,
        expected_category=TECHNOLOGY_FINANCE_REGISTRATION.name,
    )


def synchronize_scenario_mapping(
    session: Session,
    profile: ScenarioRegistration,
    settings: Settings,
) -> MappingSyncResult:
    """Read, validate, and atomically publish one scenario profile's mapping."""
    if not profile.is_executable_profile:
        raise ValueError(f"scenario {profile.id!r} has no executable mapping profile")
    source = read_mapping_source(
        profile.mapping_path(settings), expected_category=profile.name
    )
    return _synchronize_mapping(
        session,
        source,
        settings,
        scenario_id=profile.id,
        expected_category=profile.name,
    )


def _synchronize_mapping(
    session: Session,
    source: MappingSyncSource,
    settings: Settings,
    *,
    scenario_id: str,
    expected_category: str,
) -> MappingSyncResult:
    existing = session.scalar(
        select(FiveArticlesMappingVersion).where(
            FiveArticlesMappingVersion.scenario_id == scenario_id,
            FiveArticlesMappingVersion.source_hash == source.source_hash,
        )
    )
    if existing is not None:
        return MappingSyncResult(version=existing, reused=True)

    normalized_rows, normalization_errors, normalizations = _normalize_rows(
        source.rows,
        expected_category=expected_category,
    )
    catalog = load_current_catalog_facts(session, settings)
    validation_errors = [*normalization_errors, *catalog.errors]
    if catalog.version is not None:
        validation_errors.extend(_validate_against_catalog(normalized_rows, catalog))
    validation_errors.extend(_find_duplicate_rows(normalized_rows))

    version_number = session.scalar(
        select(func.max(FiveArticlesMappingVersion.version)).where(
            FiveArticlesMappingVersion.scenario_id == scenario_id
        )
    )
    version = FiveArticlesMappingVersion(
        scenario_id=scenario_id,
        version=(version_number or 0) + 1,
        source_hash=source.source_hash,
        status="draft",
        validation_report={},
    )
    session.add(version)
    session.flush()

    is_valid = bool(normalized_rows) and not validation_errors
    report: dict[str, object] = {
        "valid": is_valid,
        "scenario_id": scenario_id,
        "source_hash": source.source_hash,
        "source_row_count": len(source.rows),
        "normalized_row_count": len(normalized_rows),
        "published_row_count": len(normalized_rows) if is_valid else 0,
        "catalog_version_id": catalog.version.id if catalog.version is not None else None,
        "catalog_version": catalog.version.version if catalog.version is not None else None,
        "errors": validation_errors,
        "normalizations": normalizations,
    }
    if not normalized_rows and not validation_errors:
        report["errors"] = [
            {"type": "empty_mapping", "message": "mapping contains no data rows"}
        ]
        is_valid = False
        report["valid"] = False

    version.validation_report = report
    if not is_valid:
        version.status = "invalid"
        session.flush()
        return MappingSyncResult(version=version, reused=False)

    session.add_all(
        [
            FiveArticlesMappingRow(
                mapping_version_id=version.id,
                scenario_id=scenario_id,
                neic_code=row.neic_code,
                code_level=row.code_level,
                neic_name=row.neic_name,
                subject=row.subject,
                tier1=row.tier1,
                tier2=row.tier2,
                tier3=row.tier3,
                tier4=row.tier4,
                source_row=row.source_row,
            )
            for row in normalized_rows
        ]
    )
    session.flush()
    version.status = "published"
    session.flush()
    return MappingSyncResult(version=version, reused=False)


def load_current_catalog_facts(session: Session, settings: Settings) -> CatalogFacts:
    version = session.scalar(
        select(NationalEconomyCatalogVersion)
        .where(
            NationalEconomyCatalogVersion.embedding_model
            == settings.siliconflow_embedding_model,
            NationalEconomyCatalogVersion.embedding_dimension
            == settings.embedding_dimension,
        )
        .order_by(
            NationalEconomyCatalogVersion.created_at.desc(),
            NationalEconomyCatalogVersion.id.desc(),
        )
        .limit(1)
    )
    if version is None:
        return CatalogFacts(
            version=None,
            four_digit_names={},
            three_digit_names={},
            two_digit_names={},
            four_digit_name_codes={},
            three_digit_name_codes={},
            two_digit_name_codes={},
            errors=(
                {
                    "type": "catalog_version_not_found",
                    "embedding_model": settings.siliconflow_embedding_model,
                    "embedding_dimension": settings.embedding_dimension,
                },
            ),
        )

    statement = (
        select(
            NationalEconomyIndustryChunk.industry_code,
            NationalEconomyIndustryChunk.industry_name,
            NationalEconomyIndustryChunk.major_category_code,
            NationalEconomyIndustryChunk.major_category_name,
            NationalEconomyIndustryChunk.middle_category_code,
            NationalEconomyIndustryChunk.middle_category_name,
        )
        .where(NationalEconomyIndustryChunk.catalog_version_id == version.id)
        .distinct()
    )
    rows = session.execute(statement).all()

    four_digit_names: dict[str, set[str]] = {}
    three_digit_names: dict[str, set[str]] = {}
    two_digit_names: dict[str, set[str]] = {}
    four_digit_name_codes: dict[str, set[str]] = {}
    three_digit_name_codes: dict[str, set[str]] = {}
    two_digit_name_codes: dict[str, set[str]] = {}
    errors: list[dict[str, object]] = []
    for row in rows:
        industry_code = _normalize_catalog_industry_code(row.industry_code)
        industry_name = _normalize_name(row.industry_name)
        if industry_code is None or not industry_name:
            errors.append(
                {
                    "type": "invalid_catalog_industry",
                    "catalog_version_id": version.id,
                    "industry_code": row.industry_code,
                    "industry_name": row.industry_name,
                }
            )
        elif len(industry_code) == 4:
            four_digit_names.setdefault(industry_code, set()).add(industry_name)
            four_digit_name_codes.setdefault(industry_name, set()).add(industry_code)

        middle_code = _middle_category_digits(row.middle_category_code)
        middle_name = _normalize_name(row.middle_category_name)
        if middle_code is not None and middle_name:
            three_digit_names.setdefault(middle_code, set()).add(middle_name)
            three_digit_name_codes.setdefault(middle_name, set()).add(middle_code)

        major_code = _major_category_digits(row.major_category_code)
        major_name = _normalize_name(row.major_category_name)
        if major_code is None or not major_name:
            errors.append(
                {
                    "type": "invalid_catalog_major_category",
                    "catalog_version_id": version.id,
                    "major_category_code": row.major_category_code,
                    "major_category_name": row.major_category_name,
                }
            )
        else:
            two_digit_names.setdefault(major_code, set()).add(major_name)
            two_digit_name_codes.setdefault(major_name, set()).add(major_code)

    errors.extend(_catalog_ambiguity_errors("4", four_digit_names))
    errors.extend(_catalog_ambiguity_errors("3", three_digit_names))
    errors.extend(_catalog_ambiguity_errors("2", two_digit_names))
    return CatalogFacts(
        version=version,
        four_digit_names=_freeze_sets(four_digit_names),
        three_digit_names=_freeze_sets(three_digit_names),
        two_digit_names=_freeze_sets(two_digit_names),
        four_digit_name_codes=_freeze_sets(four_digit_name_codes),
        three_digit_name_codes=_freeze_sets(three_digit_name_codes),
        two_digit_name_codes=_freeze_sets(two_digit_name_codes),
        errors=tuple(errors),
    )


def _normalize_rows(
    rows: Sequence[RawMappingRow],
    *,
    expected_category: str,
) -> tuple[
    tuple[NormalizedMappingRow, ...],
    list[dict[str, object]],
    list[dict[str, object]],
]:
    normalized: list[NormalizedMappingRow] = []
    errors: list[dict[str, object]] = []
    normalizations: list[dict[str, object]] = []
    for row in rows:
        raw_category = row.values.get(OPTIONAL_CATEGORY_HEADER)
        category = _normalize_text(raw_category)
        if category and category != expected_category:
            errors.append(
                {
                    "type": "category_mismatch",
                    "source_row": row.source_row,
                    "expected": expected_category,
                    "actual": category,
                }
            )

        raw_code = row.values.get("国民经济行业代码")
        code = _normalize_neic_code(raw_code)
        if code is None:
            errors.append(
                {
                    "type": "invalid_code",
                    "source_row": row.source_row,
                    "value": _audit_value(raw_code),
                }
            )

        raw_name = row.values.get("国民经济行业名称")
        name = _normalize_name(raw_name)
        subject = _normalize_text(row.values.get("主题"))
        tier1 = _normalize_text(row.values.get("第一层名称"))
        tier2 = _optional_text(row.values.get("第二层名称"))
        tier3 = _optional_text(row.values.get("第三层名称"))
        tier4 = _optional_text(row.values.get("第四层名称"))

        for field, value in (
            ("NEIC_Name", name),
            ("主题", subject),
            ("第一层名称", tier1),
        ):
            if not value:
                errors.append(
                    {
                        "type": "missing_required_value",
                        "source_row": row.source_row,
                        "field": field,
                    }
                )

        if code is not None and _audit_value(raw_code) != code:
            normalizations.append(
                {
                    "source_row": row.source_row,
                    "field": "NEIC_Code",
                    "original": _audit_value(raw_code),
                    "normalized": code,
                }
            )
        raw_name_text = "" if raw_name is None else str(raw_name)
        if name and raw_name_text != name:
            normalizations.append(
                {
                    "source_row": row.source_row,
                    "field": "NEIC_Name",
                    "original": raw_name_text,
                    "normalized": name,
                }
            )

        if code is None or not name or not subject or not tier1:
            continue
        normalized.append(
            NormalizedMappingRow(
                source_row=row.source_row,
                neic_code=code,
                code_level=len(code),
                neic_name=name,
                comparison_name=name,
                subject=subject,
                tier1=tier1,
                tier2=tier2,
                tier3=tier3,
                tier4=tier4,
            )
        )
    return tuple(normalized), errors, normalizations


def _validate_against_catalog(
    rows: Sequence[NormalizedMappingRow], catalog: CatalogFacts
) -> list[dict[str, object]]:
    errors: list[dict[str, object]] = []
    for row in rows:
        if row.code_level == 4:
            names = catalog.four_digit_names
            name_codes = catalog.four_digit_name_codes
        elif row.code_level == 3:
            names = catalog.three_digit_names
            name_codes = catalog.three_digit_name_codes
        else:
            names = catalog.two_digit_names
            name_codes = catalog.two_digit_name_codes

        expected_names = names.get(row.neic_code)
        if expected_names is None:
            errors.append(
                {
                    "type": "code_not_found",
                    "source_row": row.source_row,
                    "code_level": row.code_level,
                    "neic_code": row.neic_code,
                    "neic_name": row.neic_name,
                }
            )
            continue
        if row.comparison_name in expected_names:
            continue

        matching_codes = sorted(name_codes.get(row.comparison_name, ()))
        errors.append(
            {
                "type": "name_code_conflict" if matching_codes else "name_mismatch",
                "source_row": row.source_row,
                "code_level": row.code_level,
                "neic_code": row.neic_code,
                "neic_name": row.neic_name,
                "expected_names": sorted(expected_names),
                "name_matching_codes": matching_codes,
            }
        )
    return errors


def _find_duplicate_rows(
    rows: Sequence[NormalizedMappingRow],
) -> list[dict[str, object]]:
    first_rows: dict[tuple[object, ...], int] = {}
    errors: list[dict[str, object]] = []
    for row in rows:
        first_row = first_rows.setdefault(row.duplicate_key, row.source_row)
        if first_row != row.source_row:
            errors.append(
                {
                    "type": "exact_duplicate",
                    "source_row": row.source_row,
                    "duplicate_of_source_row": first_row,
                    "neic_code": row.neic_code,
                    "neic_name": row.neic_name,
                    "taxonomy": [
                        row.subject,
                        row.tier1,
                        row.tier2,
                        row.tier3,
                        row.tier4,
                    ],
                }
            )
    return errors


def _normalize_neic_code(value: Any) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        digits = str(value)
    elif isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer():
            return None
        digits = str(int(value))
    else:
        text = unicodedata.normalize("NFKC", str(value))
        compact = re.sub(r"\s+", "", text)
        if re.fullmatch(r"\d+\.0+", compact):
            try:
                digits = str(int(Decimal(compact)))
            except (InvalidOperation, ValueError):
                return None
        else:
            digits = compact
    if not digits.isdigit():
        return None
    if len(digits) == 1:
        digits = digits.zfill(2)
    return digits if len(digits) in (2, 3, 4) else None


def _normalize_catalog_industry_code(value: Any) -> str | None:
    text = _normalize_text(value)
    match = re.fullmatch(r"[A-Za-z]?(\d{2,4})", text)
    return match.group(1) if match is not None else None


def _major_category_digits(value: Any) -> str | None:
    text = re.sub(r"\s+", "", _normalize_text(value))
    match = MAJOR_CATEGORY_CODE_PATTERN.fullmatch(text)
    return match.group(1) if match is not None else None


def _middle_category_digits(value: Any) -> str | None:
    text = re.sub(r"\s+", "", _normalize_text(value))
    match = re.fullmatch(r"[A-Za-z]?(\d{3})", text)
    return match.group(1) if match is not None else None


def _canonical_header(value: Any) -> str:
    if value is None:
        return ""
    first_line = str(value).splitlines()[0]
    return _normalize_text(first_line)


def _normalize_name(value: Any) -> str:
    return _normalize_text(value)


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    normalized = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", normalized).strip()


def _optional_text(value: Any) -> str | None:
    normalized = _normalize_text(value)
    return normalized or None


def _is_blank(value: Any) -> bool:
    return not _normalize_text(value)


def _audit_value(value: Any) -> str:
    return "" if value is None else str(value)


def _freeze_sets(values: Mapping[str, set[str]]) -> dict[str, frozenset[str]]:
    return {key: frozenset(item_values) for key, item_values in values.items()}


def _catalog_ambiguity_errors(
    code_level: str, values: Mapping[str, set[str]]
) -> list[dict[str, object]]:
    return [
        {
            "type": "ambiguous_catalog_code",
            "code_level": int(code_level),
            "neic_code": code,
            "names": sorted(names),
        }
        for code, names in values.items()
        if len(names) > 1
    ]

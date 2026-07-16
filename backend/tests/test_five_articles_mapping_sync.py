from collections.abc import Iterator, Sequence
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
import pytest
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.db.session import get_sessionmaker
from app.models import (
    FiveArticlesMappingRow,
    FiveArticlesMappingVersion,
    NationalEconomyCatalogVersion,
    NationalEconomyIndustryChunk,
)
from app.services.scenario_registry import (
    DIGITAL_FINANCE_REGISTRATION,
    GREEN_FINANCE_REGISTRATION,
    PENSION_FINANCE_REGISTRATION,
    TECHNOLOGY_FINANCE_REGISTRATION,
    ScenarioRegistration,
)
from app.services.technology_finance_mapping_sync import (
    _normalize_neic_code,
    synchronize_scenario_mapping,
)


SCENARIO_PROFILES = (
    TECHNOLOGY_FINANCE_REGISTRATION,
    DIGITAL_FINANCE_REGISTRATION,
    PENSION_FINANCE_REGISTRATION,
)
REQUIRED_HEADERS = (
    "主题\nSubject",
    "第一层名称\nTier1_Name",
    "第二层名称\nTier2_Name",
    "第三层名称\nTier3_Name",
    "第四层名称\nTier4_Name",
    "国民经济行业代码\nNEIC_Code",
    "国民经济行业名称\nNEIC_Name",
)


def _mapping_row(
    code: object,
    name: str,
    *,
    subject: str = "场景主题",
    tier1: str = "场景一级标签",
) -> tuple[object, ...]:
    return (subject, tier1, None, None, None, code, name)


def _write_mapping(
    path: Path,
    rows: Sequence[Sequence[object]],
    *,
    category: str | None = None,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    if category is None:
        worksheet.append(REQUIRED_HEADERS)
        for row in rows:
            worksheet.append(tuple(row))
    else:
        worksheet.append(("属于类别", *REQUIRED_HEADERS))
        for row in rows:
            worksheet.append((category, *row))
    workbook.save(path)
    workbook.close()


def _write_merged_mapping(
    path: Path,
    rows: Sequence[tuple[str, Sequence[object]]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(("属于类别", *REQUIRED_HEADERS))
    for category, row in rows:
        worksheet.append((category, *row))
    workbook.save(path)
    workbook.close()


@pytest.fixture()
def scenario_mapping_context() -> Iterator[tuple[Session, str]]:
    session = get_sessionmaker()()
    embedding_model = f"multi-scenario-mapping-{uuid4().hex}"
    catalog_version = NationalEconomyCatalogVersion(
        version=f"catalog-{uuid4().hex}",
        source_hash=uuid4().hex * 2,
        embedding_model=embedding_model,
        embedding_dimension=4096,
    )
    session.add(catalog_version)
    session.flush()
    session.add_all(
        [
            NationalEconomyIndustryChunk(
                catalog_version_id=catalog_version.id,
                major_category_code="C27",
                major_category_name="医药制造业",
                middle_category_code="271",
                middle_category_name="化学药品制造",
                industry_code="2710",
                industry_name="化学药品原料药制造",
                source_row=2,
                text="化学药品原料药制造定义",
                chunk_type="definition",
                embedding=[0.0] * 4096,
            ),
            NationalEconomyIndustryChunk(
                catalog_version_id=catalog_version.id,
                major_category_code="C27",
                major_category_name="医药制造业",
                middle_category_code="272",
                middle_category_name="化学药品制剂制造",
                industry_code="2720",
                industry_name="化学药品制剂制造",
                source_row=3,
                text="化学药品制剂制造定义",
                chunk_type="definition",
                embedding=[0.0] * 4096,
            ),
        ]
    )
    session.flush()
    try:
        yield session, embedding_model
    finally:
        session.rollback()
        session.close()


def _settings_for(
    profile: ScenarioRegistration,
    path: Path,
    embedding_model: str,
) -> Settings:
    assert profile.uses_five_articles_mapping is True
    values = dict(
        _env_file=None,
        SILICONFLOW_EMBEDDING_MODEL=embedding_model,
        EMBEDDING_DIMENSION=4096,
        FIVE_ARTICLES_MAPPING_SOURCE_PATH=path,
    )
    if profile is GREEN_FINANCE_REGISTRATION:
        values["GREEN_FINANCE_MAPPING_SOURCE_PATH"] = path
        values["SILICONFLOW_API_KEY"] = "test-key"
    return Settings(**values)


GREEN_HEADERS = (
    "属于类别",
    "主题\nSubject",
    "第一层名称\nTier1_Name",
    "第二层名称\nTier2_Name",
    "条件/标准",
    "国民经济行业代码\nNEIC_Code",
    "国民经济行业名称\nNEIC_Name",
)


def _write_green_mapping(path: Path, rows: Sequence[Sequence[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(GREEN_HEADERS)
    for row in rows:
        worksheet.append(tuple(row))
    workbook.save(path)
    workbook.close()


def test_green_mapping_publishes_condition_embeddings_and_placeholder_rows(
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / "green.xlsx"
    _write_green_mapping(
        path,
        (
            ("绿色金融", "绿色主题", "绿色一级", "绿色二级", "节能锅炉", "2710", "化学药品原料药制造"),
            ("绿色金融", "绿色主题", "绿色一级", "绿色二级", "绿色项目", "-", "无行业代码"),
        ),
    )
    monkeypatch.setattr(
        "app.services.technology_finance_mapping_sync.embed_texts",
        lambda texts, settings: tuple((0.1,) * settings.embedding_dimension for _ in texts),
    )

    result = synchronize_scenario_mapping(
        session,
        GREEN_FINANCE_REGISTRATION,
        _settings_for(GREEN_FINANCE_REGISTRATION, path, embedding_model),
    )

    assert result.version.status == "published"
    rows = session.scalars(
        select(FiveArticlesMappingRow)
        .where(FiveArticlesMappingRow.mapping_version_id == result.version.id)
        .order_by(FiveArticlesMappingRow.source_row)
    ).all()
    assert [(row.neic_code, row.code_level) for row in rows] == [("2710", 4), ("-", None)]
    assert all(row.condition_criteria for row in rows)
    assert all(
        row.condition_embedding is not None and len(row.condition_embedding) == 4096
        for row in rows
    )


def test_green_mapping_identity_distinguishes_conditions_but_rejects_exact_duplicates(
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, embedding_model = scenario_mapping_context
    monkeypatch.setattr(
        "app.services.technology_finance_mapping_sync.embed_texts",
        lambda texts, settings: tuple(
            (0.1,) * settings.embedding_dimension for _ in texts
        ),
    )
    distinct_path = tmp_path / "green-distinct-conditions.xlsx"
    shared = (
        "绿色金融",
        "能源绿色低碳转型",
        "新能源与清洁能源装备制造",
        "新型储能产品制造",
    )
    _write_green_mapping(
        distinct_path,
        (
            (*shared, "超级电容储能产品制造", "2710", "化学药品原料药制造"),
            (*shared, "储能电池制造", "2710", "化学药品原料药制造"),
        ),
    )

    distinct = synchronize_scenario_mapping(
        session,
        GREEN_FINANCE_REGISTRATION,
        _settings_for(GREEN_FINANCE_REGISTRATION, distinct_path, embedding_model),
    )

    assert distinct.version.status == "published"
    assert distinct.version.validation_report["published_row_count"] == 2

    duplicate_path = tmp_path / "green-exact-duplicate.xlsx"
    duplicate_row = (
        *shared,
        "储能电池制造",
        "2710",
        "化学药品原料药制造",
    )
    _write_green_mapping(duplicate_path, (duplicate_row, duplicate_row))

    duplicate = synchronize_scenario_mapping(
        session,
        GREEN_FINANCE_REGISTRATION,
        _settings_for(GREEN_FINANCE_REGISTRATION, duplicate_path, embedding_model),
    )

    assert duplicate.version.status == "invalid"
    assert any(
        error["type"] == "exact_duplicate"
        for error in duplicate.version.validation_report["errors"]
    )


def test_green_mapping_embedding_failure_does_not_publish_a_partial_version(
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / "green-failure.xlsx"
    _write_green_mapping(
        path,
        (("绿色金融", "绿色主题", "绿色一级", "绿色二级", "节能锅炉", "2710", "化学药品原料药制造"),),
    )
    monkeypatch.setattr(
        "app.services.technology_finance_mapping_sync.embed_texts",
        lambda *_: (_ for _ in ()).throw(RuntimeError("embedding unavailable")),
    )

    with pytest.raises(RuntimeError, match="embedding unavailable"):
        synchronize_scenario_mapping(
            session,
            GREEN_FINANCE_REGISTRATION,
            _settings_for(GREEN_FINANCE_REGISTRATION, path, embedding_model),
        )
    assert session.scalar(
        select(func.count(FiveArticlesMappingVersion.id)).where(
            FiveArticlesMappingVersion.scenario_id == GREEN_FINANCE_REGISTRATION.id
        )
    ) == 0


def test_non_green_mapping_rejects_condition_header_and_tiers_above_depth(
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / "digital-condition.xlsx"
    _write_mapping(path, (_mapping_row("2710", "化学药品原料药制造"),), category="数字金融")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(("属于类别", *REQUIRED_HEADERS, "条件/标准"))
    worksheet.append(("数字金融", *_mapping_row("2710", "化学药品原料药制造"), "条件"))
    workbook.save(path)
    workbook.close()
    with pytest.raises(ValueError, match="条件/标准"):
        synchronize_scenario_mapping(session, DIGITAL_FINANCE_REGISTRATION, _settings_for(DIGITAL_FINANCE_REGISTRATION, path, embedding_model))

    path = tmp_path / "pension-tier4.xlsx"
    _write_mapping(
        path,
        (
            (
                "场景主题",
                "场景一级标签",
                None,
                None,
                "不允许的第四层",
                "2710",
                "化学药品原料药制造",
            ),
        ),
        category="养老金融",
    )
    result = synchronize_scenario_mapping(session, PENSION_FINANCE_REGISTRATION, _settings_for(PENSION_FINANCE_REGISTRATION, path, embedding_model))
    assert result.version.status == "invalid"
    assert result.version.validation_report["errors"][0]["type"] == "tier_exceeds_declared_depth"


@pytest.mark.parametrize("profile", SCENARIO_PROFILES, ids=lambda item: item.id)
def test_scenario_profile_mapping_publishes_valid_two_and_four_digit_rows(
    profile: ScenarioRegistration,
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / f"{profile.id}.xlsx"
    _write_mapping(
        path,
        (
            _mapping_row("27\u3000", " 医药制造业 "),
            _mapping_row(2710.0, "化学药品原料药制造", subject="另一主题"),
        ),
        category=profile.name,
    )

    result = synchronize_scenario_mapping(
        session, profile, _settings_for(profile, path, embedding_model)
    )

    assert result.reused is False
    assert result.version.scenario_id == profile.id
    assert result.version.status == "published"
    assert result.version.validation_report["valid"] is True
    assert result.version.validation_report["published_row_count"] == 2
    rows = session.scalars(
        select(FiveArticlesMappingRow)
        .where(FiveArticlesMappingRow.mapping_version_id == result.version.id)
        .order_by(FiveArticlesMappingRow.source_row)
    ).all()
    assert [(row.neic_code, row.code_level) for row in rows] == [("27", 2), ("2710", 4)]
    assert {row.scenario_id for row in rows} == {profile.id}


@pytest.mark.parametrize("profile", SCENARIO_PROFILES, ids=lambda item: item.id)
def test_scenario_profile_mapping_reuses_the_same_source_hash(
    profile: ScenarioRegistration,
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / f"{profile.id}-reuse.xlsx"
    _write_mapping(
        path, (_mapping_row("2710", "化学药品原料药制造"),), category=profile.name
    )
    settings = _settings_for(profile, path, embedding_model)

    first = synchronize_scenario_mapping(session, profile, settings)
    second = synchronize_scenario_mapping(session, profile, settings)

    assert first.reused is False
    assert second.reused is True
    assert second.version.id == first.version.id
    assert session.scalar(
        select(func.count(FiveArticlesMappingVersion.id)).where(
            FiveArticlesMappingVersion.scenario_id == profile.id,
            FiveArticlesMappingVersion.source_hash == first.version.source_hash,
        )
    ) == 1


@pytest.mark.parametrize("profile", SCENARIO_PROFILES, ids=lambda item: item.id)
def test_scenario_profile_mapping_rejects_a_nonexistent_code(
    profile: ScenarioRegistration,
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / f"{profile.id}-missing-code.xlsx"
    _write_mapping(path, (_mapping_row("9999", "不存在行业"),), category=profile.name)

    result = synchronize_scenario_mapping(
        session, profile, _settings_for(profile, path, embedding_model)
    )

    assert result.version.status == "invalid"
    assert result.version.validation_report["errors"][0]["type"] == "code_not_found"
    assert result.version.validation_report["published_row_count"] == 0


@pytest.mark.parametrize("profile", SCENARIO_PROFILES, ids=lambda item: item.id)
def test_scenario_profile_mapping_rejects_a_name_code_conflict(
    profile: ScenarioRegistration,
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / f"{profile.id}-name-conflict.xlsx"
    _write_mapping(
        path, (_mapping_row("2710", "化学药品制剂制造"),), category=profile.name
    )

    result = synchronize_scenario_mapping(
        session, profile, _settings_for(profile, path, embedding_model)
    )

    error = result.version.validation_report["errors"][0]
    assert result.version.status == "invalid"
    assert error["type"] == "name_code_conflict"
    assert error["name_matching_codes"] == ["2720"]


@pytest.mark.parametrize("profile", SCENARIO_PROFILES, ids=lambda item: item.id)
def test_scenario_profile_mapping_rejects_exact_duplicates(
    profile: ScenarioRegistration,
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / f"{profile.id}-duplicate.xlsx"
    _write_mapping(
        path,
        (
            _mapping_row("2710", "化学药品原料药制造"),
            _mapping_row("2710.0", "化学药品原料药制造"),
        ),
        category=profile.name,
    )

    result = synchronize_scenario_mapping(
        session, profile, _settings_for(profile, path, embedding_model)
    )

    errors = result.version.validation_report["errors"]
    assert result.version.status == "invalid"
    assert next(error for error in errors if error["type"] == "exact_duplicate") == {
        "type": "exact_duplicate",
        "source_row": 3,
        "duplicate_of_source_row": 2,
        "neic_code": "2710",
        "neic_name": "化学药品原料药制造",
        "taxonomy": ["场景主题", "场景一级标签", None, None, None],
    }


@pytest.mark.parametrize("profile", SCENARIO_PROFILES, ids=lambda item: item.id)
def test_scenario_profile_mapping_ignores_rows_from_other_categories(
    profile: ScenarioRegistration,
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / f"{profile.id}-category-mismatch.xlsx"
    _write_mapping(
        path,
        (_mapping_row("2710", "化学药品原料药制造"),),
        category="其他金融",
    )

    result = synchronize_scenario_mapping(
        session, profile, _settings_for(profile, path, embedding_model)
    )

    assert result.version.status == "invalid"
    assert result.version.validation_report["published_row_count"] == 0
    assert result.version.validation_report["errors"] == [
        {"type": "empty_mapping", "message": "mapping contains no data rows"}
    ]
    assert session.scalar(
        select(func.count(FiveArticlesMappingRow.id)).where(
            FiveArticlesMappingRow.mapping_version_id == result.version.id
        )
    ) == 0


def test_unrelated_category_change_reuses_current_scenario_version(
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / "merged.xlsx"
    _write_merged_mapping(
        path,
        (
            ("数字金融", _mapping_row("2710", "化学药品原料药制造")),
            ("绿色金融", _mapping_row("2720", "化学药品制剂制造")),
        ),
    )
    settings = _settings_for(DIGITAL_FINANCE_REGISTRATION, path, embedding_model)

    first = synchronize_scenario_mapping(session, DIGITAL_FINANCE_REGISTRATION, settings)
    _write_merged_mapping(
        path,
        (
            ("数字金融", _mapping_row("2710", "化学药品原料药制造")),
            ("绿色金融", _mapping_row("2720", "化学药品制剂制造", subject="已修改")),
        ),
    )
    second = synchronize_scenario_mapping(session, DIGITAL_FINANCE_REGISTRATION, settings)

    assert first.version.status == "published"
    assert second.reused is True
    assert second.version.id == first.version.id


def test_scenario_profile_mapping_publishes_valid_three_digit_middle_class_row(
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / "middle-class.xlsx"
    _write_mapping(
        path,
        (_mapping_row("271", "化学药品制造"),),
        category=PENSION_FINANCE_REGISTRATION.name,
    )

    result = synchronize_scenario_mapping(
        session,
        PENSION_FINANCE_REGISTRATION,
        _settings_for(PENSION_FINANCE_REGISTRATION, path, embedding_model),
    )

    assert result.version.status == "published"
    assert result.version.validation_report["published_row_count"] == 1
    row = session.scalar(
        select(FiveArticlesMappingRow).where(
            FiveArticlesMappingRow.mapping_version_id == result.version.id
        )
    )
    assert row is not None
    assert (row.neic_code, row.code_level, row.neic_name) == ("271", 3, "化学药品制造")


@pytest.mark.parametrize(
    ("raw_code", "expected"),
    (("A12", "12"), ("B123", "123"), ("Z1234", "1234")),
)
def test_mapping_code_normalization_accepts_any_v2_letter_prefix(
    raw_code: str, expected: str
) -> None:
    assert _normalize_neic_code(raw_code) == expected


def test_mapping_publishes_a_prefixed_v2_code_and_canonicalizes_a_business_name(
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    catalog = session.scalar(select(NationalEconomyCatalogVersion))
    assert catalog is not None
    session.add(
        NationalEconomyIndustryChunk(
            catalog_version_id=catalog.id,
            major_category_code="R87",
            major_category_name="广播、电视、电影和影视录音制作业",
            middle_category_code="R871",
            middle_category_name="广播",
            industry_code="8710",
            industry_name="广播",
            source_row=4,
            text="广播目录事实",
            chunk_type="catalog",
            embedding=[0.0] * 4096,
        )
    )
    session.flush()
    path = tmp_path / "canonicalized-name.xlsx"
    _write_mapping(
        path,
        (_mapping_row("R8710", "互联网广播"),),
        category=TECHNOLOGY_FINANCE_REGISTRATION.name,
    )

    result = synchronize_scenario_mapping(
        session,
        TECHNOLOGY_FINANCE_REGISTRATION,
        _settings_for(TECHNOLOGY_FINANCE_REGISTRATION, path, embedding_model),
    )

    assert result.version.status == "published"
    row = session.scalar(
        select(FiveArticlesMappingRow).where(
            FiveArticlesMappingRow.mapping_version_id == result.version.id
        )
    )
    assert row is not None
    assert (row.neic_code, row.neic_name) == ("8710", "广播")
    assert result.version.validation_report["normalizations"][-1] == {
        "type": "catalog_name_normalized",
        "source_row": 2,
        "field": "NEIC_Name",
        "original": "互联网广播",
        "normalized": "广播",
        "neic_code": "8710",
    }


def test_invalid_same_source_hash_is_revalidated_after_catalog_repair(
    tmp_path: Path,
    scenario_mapping_context: tuple[Session, str],
) -> None:
    session, embedding_model = scenario_mapping_context
    path = tmp_path / "retry-invalid-source.xlsx"
    _write_mapping(
        path,
        (_mapping_row("9999", "新增目录行业"),),
        category=TECHNOLOGY_FINANCE_REGISTRATION.name,
    )
    settings = _settings_for(TECHNOLOGY_FINANCE_REGISTRATION, path, embedding_model)
    first = synchronize_scenario_mapping(session, TECHNOLOGY_FINANCE_REGISTRATION, settings)
    assert first.version.status == "invalid"

    catalog = session.scalar(select(NationalEconomyCatalogVersion))
    assert catalog is not None
    session.add(
        NationalEconomyIndustryChunk(
            catalog_version_id=catalog.id,
            major_category_code="Z99",
            major_category_name="新增大类",
            middle_category_code="Z999",
            middle_category_name="新增中类",
            industry_code="9999",
            industry_name="新增目录行业",
            source_row=5,
            text="新增目录事实",
            chunk_type="catalog",
            embedding=[0.0] * 4096,
        )
    )
    session.flush()

    second = synchronize_scenario_mapping(session, TECHNOLOGY_FINANCE_REGISTRATION, settings)

    assert second.reused is False
    assert second.version.id == first.version.id
    assert second.version.status == "published"

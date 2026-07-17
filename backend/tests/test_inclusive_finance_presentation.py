from io import BytesIO

from openpyxl import load_workbook

from app.models import InclusiveFinanceResult, NationalEconomyClassificationCase
from app.services.national_economy_case_export import export_case_workbook


def test_inclusive_finance_export_uses_chinese_borrower_and_evidence_labels() -> None:
    case = NationalEconomyClassificationCase(
        id=1,
        scenario="inclusive_finance",
        original_filename="普惠测试.docx",
        input_payload={},
        status="completed",
    )
    result = InclusiveFinanceResult(
        id=1,
        case=case,
        scenario_id="inclusive_finance",
        version=1,
        status="completed",
        stage_a_result_id=1,
        borrower_type="farmer",
        is_operating_loan=True,
        credit_amount_wan=500,
        qualifies=True,
        inclusive_category="农户经营性贷款",
        basis="测试依据",
        evidence_refs=[
            {
                "type": "field",
                "field_key": "credit_approval_opinion",
                "raw_value": "支持生产经营周转",
            }
        ],
        anomalies=[],
        determination={
            "borrower_type_basis": "农户条件命中：长期居住在乡镇",
            "structured_credit_amount_wan": 500,
            "approval_credit_amounts_wan": [500],
            "credit_amount_source": "structured_and_approval_consistent",
            "farmer_registration_address_support": "注册地址包含乡，可作为辅助佐证",
        },
    )

    workbook = load_workbook(
        BytesIO(export_case_workbook(case, inclusive_finance_results=[result]))
    )
    sheet = workbook["普惠金融判定"]
    headers = [cell.value for cell in sheet[1]]
    row = dict(zip(headers, next(sheet.iter_rows(min_row=2, values_only=True)), strict=True))

    assert row["借款主体"] == "农户"
    assert row["主体条件"] == "农户条件命中：长期居住在乡镇"
    assert row["结构化授信额度(万元)"] == 500
    assert row["审批意见批复额度(万元)"] == "500"
    assert row["额度来源及一致性"] == "结构化额度与审批意见一致"
    assert row["注册地址辅助"] == "注册地址包含乡，可作为辅助佐证"
    assert row["业务证据摘要"] == "授信审批意见：支持生产经营周转"

from collections.abc import Iterator
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app.api.routes import national_economy as route_module
from app.db.session import get_db, get_sessionmaker
from app.main import app
from app.models import (
    InclusiveFinanceResult,
    NationalEconomyClassificationCase,
    NationalEconomyClassificationResult,
)
from app.services.inclusive_finance_workflow import (
    InclusiveFinanceWorkflowResult,
    classify_inclusive_finance_case,
    reclassify_inclusive_finance_case,
    run_inclusive_finance_stage_b,
)
from app.services.scenario_workflow_handlers import ScenarioWorkflowHandler


SCENARIO_ID = "inclusive_finance"


@pytest.fixture()
def workflow_context() -> Iterator[tuple[Session, NationalEconomyClassificationCase]]:
    session = get_sessionmaker()()
    case = NationalEconomyClassificationCase(
        scenario=SCENARIO_ID,
        original_filename="inclusive.docx",
        input_payload={"entity_type": "小型企业", "credit_amount": "500万元"},
        status="pending_classification",
    )
    session.add(case)
    session.commit()
    try:
        yield session, case
    finally:
        session.rollback()
        session.execute(delete(InclusiveFinanceResult).where(InclusiveFinanceResult.case_id == case.id))
        session.execute(delete(NationalEconomyClassificationResult).where(NationalEconomyClassificationResult.case_id == case.id))
        session.execute(delete(NationalEconomyClassificationCase).where(NationalEconomyClassificationCase.id == case.id))
        session.commit()
        session.close()


def _stage_a(session: Session, case: NationalEconomyClassificationCase, *, version: int, status: str = "completed") -> NationalEconomyClassificationResult:
    result = NationalEconomyClassificationResult(
        case_id=case.id, version=version, status=status,
        industry_code="5210" if status == "completed" else None,
        industry_major_code="F52" if status == "completed" else None,
        industry_name="综合零售" if status == "completed" else None,
        rationale="test", candidate_snapshot=[], model_output={"stage": "a"},
    )
    session.add(result); case.status = status; session.commit(); session.refresh(result)
    return result


def _decision(*, status: str = "completed") -> dict[str, object]:
    return {
        "status": status, "borrower_type": "enterprise", "computed_size": "小型",
        "filled_size": "小型", "qualifies": status == "completed",
        "inclusive_category": "小微企业贷款" if status == "completed" else None,
        "basis": "deterministic test", "evidence_refs": [], "anomalies": [],
        "determination": {
            "size_consistent": True,
            "borrower_type_basis": "主体类型填报：小型企业",
            "structured_credit_amount_wan": 500.0,
            "approval_credit_amounts_wan": [500.0],
            "credit_amount_source": "structured_and_approval_consistent",
            "credit_amount_consistent": True,
            "credit_amount_conflict": False,
            "farmer_registration_address_support": "注册地址未形成农户身份依据",
        }, "is_operating_loan": True,
        "credit_amount_wan": 500.0,
    }


def test_incomplete_stage_a_short_circuits_without_inclusive_result(workflow_context: tuple[Session, NationalEconomyClassificationCase]) -> None:
    session, case = workflow_context
    stage_a = _stage_a(session, case, version=1, status="needs_review")

    outcome = run_inclusive_finance_stage_b(session, case, stage_a)

    assert outcome.stage_a_result.id == stage_a.id
    assert outcome.stage_b_result is None
    assert session.scalar(select(func.count(InclusiveFinanceResult.id)).where(InclusiveFinanceResult.case_id == case.id)) == 0


def test_completed_stage_a_is_idempotent_and_failure_preserves_stage_a(workflow_context: tuple[Session, NationalEconomyClassificationCase]) -> None:
    session, case = workflow_context
    stage_a = _stage_a(session, case, version=1)

    failed = run_inclusive_finance_stage_b(session, case, stage_a, determiner=MagicMock(side_effect=RuntimeError("boom")))
    assert failed.stage_b_result is not None and failed.stage_b_result.status == "classification_failed"
    assert session.get(NationalEconomyClassificationResult, stage_a.id) is not None

    completed = run_inclusive_finance_stage_b(session, case, stage_a, determiner=MagicMock(return_value=_decision()))
    duplicate = run_inclusive_finance_stage_b(session, case, stage_a, determiner=MagicMock(return_value=_decision()))
    assert completed.stage_b_result is not None and completed.stage_b_result.status == "completed"
    assert duplicate.stage_b_result is not None and duplicate.stage_b_result.id == completed.stage_b_result.id
    assert [item.status for item in session.scalars(select(InclusiveFinanceResult).where(InclusiveFinanceResult.case_id == case.id).order_by(InclusiveFinanceResult.version)).all()] == ["classification_failed", "completed"]


def test_objection_uses_new_stage_a_and_new_inclusive_version(workflow_context: tuple[Session, NationalEconomyClassificationCase]) -> None:
    session, case = workflow_context
    first = _stage_a(session, case, version=1)
    first_outcome = run_inclusive_finance_stage_b(session, case, first, determiner=MagicMock(return_value=_decision(status="not_applicable")))
    second = _stage_a(session, case, version=2)
    reclassified = reclassify_inclusive_finance_case(session, case, "异议", stage_a_reclassifier=MagicMock(return_value=second), determiner=MagicMock(return_value=_decision()))
    assert first_outcome.stage_b_result is not None
    assert reclassified.stage_b_result is not None
    assert reclassified.stage_b_result.stage_a_result_id == second.id
    assert reclassified.stage_b_result.version == 2


@pytest.fixture()
def client(workflow_context: tuple[Session, NationalEconomyClassificationCase]) -> Iterator[TestClient]:
    session, _ = workflow_context
    def override_get_db():
        yield session
    app.dependency_overrides[get_db] = override_get_db
    try:
        with TestClient(app) as test_client:
            yield test_client
    finally:
        app.dependency_overrides.pop(get_db, None)


def test_inclusive_seven_endpoints_export_and_scenario_mismatch(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    template = Path("模板文件/五篇大文章/普惠金融模版.docx")
    assert client.get(f"/api/v1/scenarios/{SCENARIO_ID}/template").status_code == 200
    uploaded = client.post(f"/api/v1/scenarios/{SCENARIO_ID}/cases", files={"file": (template.name, template.read_bytes(), route_module.DOCX_MIME)})
    assert uploaded.status_code == 201
    case_id = uploaded.json()["id"]
    assert client.get(f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}").status_code == 200

    def classify(session: Session, case: NationalEconomyClassificationCase):
        stage_a = _stage_a(session, case, version=1)
        return run_inclusive_finance_stage_b(session, case, stage_a, determiner=MagicMock(return_value=_decision()))
    def reclassify(session: Session, case: NationalEconomyClassificationCase, _: str):
        stage_a = _stage_a(session, case, version=2)
        return run_inclusive_finance_stage_b(session, case, stage_a, determiner=MagicMock(return_value=_decision(status="needs_review")))
    handler = ScenarioWorkflowHandler(classify, reclassify, InclusiveFinanceResult)
    monkeypatch.setattr(route_module, "get_scenario_workflow_handler", lambda _: handler)

    classified = client.post(f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}/classifications")
    assert classified.status_code == 200 and classified.json()["stage_b"]["status"] == "completed"
    assert classified.json()["stage_b"]["determination"]["credit_amount_source"] == "structured_and_approval_consistent"
    objection = client.post(f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}/objections", json={"objection_text": "需要复核"})
    assert objection.status_code == 200 and objection.json()["stage_b"]["status"] == "needs_review"
    history = client.get(f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}/history")
    assert [item["status"] for item in history.json()["items"]] == ["completed", "needs_review"]
    assert history.json()["items"][0]["determination"]["approval_credit_amounts_wan"] == [500.0]
    exported = client.get(f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}/export")
    workbook = load_workbook(BytesIO(exported.content))
    assert workbook.sheetnames == ["案例输入", "当前结论", "判定历史", "普惠金融判定"]
    headers = [cell.value for cell in workbook["普惠金融判定"][1]]
    assert "审批意见批复额度(万元)" in headers
    assert "额度来源及一致性" in headers
    assert client.get(f"/api/v1/scenarios/{SCENARIO_ID}/cases/999999").status_code == 404

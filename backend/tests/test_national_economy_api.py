from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app.api.routes import national_economy as route_module
from app.core.config import get_settings
from app.db.session import get_db, get_sessionmaker
from app.main import app
from app.models import (
    NationalEconomyClassificationCase,
    NationalEconomyClassificationResult,
)
from app.services.national_economy_case_ingestion import FIELD_LABELS, SCENARIO


FIXTURES = Path(__file__).parent / "fixtures" / "national_economy"


@pytest.fixture()
def db_session() -> Session:
    session = get_sessionmaker()()
    session.execute(delete(NationalEconomyClassificationResult))
    session.execute(delete(NationalEconomyClassificationCase))
    session.commit()
    try:
        yield session
    finally:
        session.rollback()
        session.execute(delete(NationalEconomyClassificationResult))
        session.execute(delete(NationalEconomyClassificationCase))
        session.commit()
        session.close()


@pytest.fixture()
def client(db_session: Session) -> TestClient:
    def override_get_db():
        yield db_session

    app.dependency_overrides[get_db] = override_get_db
    try:
        with TestClient(app) as test_client:
            yield test_client
    finally:
        app.dependency_overrides.pop(get_db, None)


def _upload(client: TestClient, fixture_name: str = "valid.docx"):
    path = FIXTURES / fixture_name
    return client.post(
        "/api/v1/national-economy/cases",
        files={"file": (path.name, path.read_bytes(), route_module.DOCX_MIME)},
    )


def _completed_result(case: NationalEconomyClassificationCase, version: int = 1):
    return NationalEconomyClassificationResult(
        case=case,
        version=version,
        status="completed",
        industry_code="0111",
        industry_name="稻谷种植",
        confidence=92,
        rationale="主营业务与目录定义一致",
        ai_summary="企业主要从事稻谷种植",
        candidate_snapshot=[{"industry_code": "0111"}],
        objection=None,
        model_output={"no_match": False},
    )


def _loan_result(
    case: NationalEconomyClassificationCase,
    *,
    version: int = 1,
    loan_industry_code: str,
    loan_industry_name: str,
    loan_matching_basis: str,
    loan_matches_enterprise: bool,
) -> NationalEconomyClassificationResult:
    result = _completed_result(case, version)
    result.loan_industry_code = loan_industry_code
    result.loan_industry_name = loan_industry_name
    result.loan_matching_basis = loan_matching_basis
    result.loan_matches_enterprise = loan_matches_enterprise
    return result


def test_scenarios_list_available_and_coming_soon_entries(client: TestClient) -> None:
    response = client.get("/api/v1/scenarios")

    assert response.status_code == 200
    scenarios = {item["id"]: item for item in response.json()["items"]}
    assert scenarios[SCENARIO]["status"] == "available"
    assert scenarios["agriculture_related"]["status"] == "coming_soon"
    assert scenarios["five_major_articles"]["status"] == "coming_soon"
    for scenario_id in (
        "technology_finance",
        "green_finance",
        "pension_finance",
        "digital_finance",
    ):
        assert scenarios[scenario_id]["status"] == "coming_soon"
        assert scenarios[scenario_id]["parent_id"] == "five_major_articles"


def test_template_download_returns_original_docx(client: TestClient) -> None:
    response = client.get("/api/v1/scenarios/national-economy/template")

    assert response.status_code == 200
    assert response.content == get_settings().national_economy_template_path.read_bytes()
    assert response.headers["content-type"] == route_module.DOCX_MIME
    assert "attachment" in response.headers["content-disposition"]


def test_upload_creates_case_and_query_returns_thirteen_fields(client: TestClient) -> None:
    upload_response = _upload(client)

    assert upload_response.status_code == 201
    created = upload_response.json()
    assert created["scenario"] == SCENARIO
    assert created["status"] == "pending_classification"

    response = client.get(f"/api/v1/national-economy/cases/{created['id']}")
    assert response.status_code == 200
    payload = response.json()
    assert payload["scenario"] == SCENARIO
    assert payload["status"] == "pending_classification"
    assert payload["current_result"] is None
    assert len(payload["input_fields"]) == 13
    assert {item["label"] for item in payload["input_fields"]} == set(FIELD_LABELS.values())


@pytest.mark.parametrize("fixture_name", ["missing_label.docx", "duplicate_label.docx", "unrecognized_label.docx"])
def test_invalid_upload_returns_actionable_issues_without_creating_case(
    client: TestClient,
    db_session: Session,
    fixture_name: str,
) -> None:
    before = db_session.scalar(select(func.count(NationalEconomyClassificationCase.id)))

    response = _upload(client, fixture_name)

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert {"message", "missing", "duplicate", "unrecognized"} <= detail.keys()
    assert any((detail["missing"], detail["duplicate"], detail["unrecognized"]))
    db_session.expire_all()
    after = db_session.scalar(select(func.count(NationalEconomyClassificationCase.id)))
    assert after == before


def test_classification_objection_history_and_failure_responses(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case_id = _upload(client).json()["id"]

    def fake_classify(session: Session, case: NationalEconomyClassificationCase):
        result = _completed_result(case)
        session.add(result)
        case.status = "completed"
        session.commit()
        session.refresh(result)
        return result

    def fake_reclassify(
        session: Session,
        case: NationalEconomyClassificationCase,
        objection_text: str,
    ):
        result = _completed_result(case, version=2)
        result.industry_code = "0112"
        result.industry_name = "小麦种植"
        result.objection = {"description": objection_text}
        session.add(result)
        session.commit()
        session.refresh(result)
        return result

    monkeypatch.setattr(route_module, "classify_case", fake_classify)
    response = client.post(f"/api/v1/national-economy/cases/{case_id}/classifications")
    assert response.status_code == 200
    classification_payload = response.json()
    assert classification_payload["version"] == 1
    assert classification_payload["industry_code"] == "0111"
    assert classification_payload["industry_name"] == "稻谷种植"
    assert classification_payload["matching_basis"] == "主营业务与目录定义一致"
    assert "confidence" not in classification_payload
    assert "ai_summary" not in classification_payload
    assert "rationale" not in classification_payload

    blank = client.post(
        f"/api/v1/national-economy/cases/{case_id}/objections",
        json={"objection_text": "   "},
    )
    assert blank.status_code == 422

    monkeypatch.setattr(route_module, "reclassify_case", fake_reclassify)
    objection = client.post(
        f"/api/v1/national-economy/cases/{case_id}/objections",
        json={"objection_text": "主营收入主要来自小麦"},
    )
    assert objection.status_code == 200
    objection_payload = objection.json()
    assert objection_payload["version"] == 2
    assert objection_payload["objection"]["description"] == "主营收入主要来自小麦"
    assert "confidence" not in objection_payload
    assert "ai_summary" not in objection_payload

    history = client.get(f"/api/v1/national-economy/cases/{case_id}/history")
    assert history.status_code == 200
    history_items = history.json()["items"]
    assert [item["version"] for item in history_items] == [1, 2]
    assert all("matching_basis" in item for item in history_items)
    assert all("confidence" not in item for item in history_items)
    assert all("ai_summary" not in item for item in history_items)

    case_payload = client.get(f"/api/v1/national-economy/cases/{case_id}").json()
    assert case_payload["current_result"]["matching_basis"] == (
        "主营业务与目录定义一致"
    )
    assert "confidence" not in case_payload["current_result"]
    assert "ai_summary" not in case_payload["current_result"]

    def failing_classify(session: Session, case: NationalEconomyClassificationCase):
        raise RuntimeError("cloud unavailable")

    monkeypatch.setattr(route_module, "classify_case", failing_classify)
    failed = client.post(f"/api/v1/national-economy/cases/{case_id}/classifications")
    assert failed.status_code == 502
    assert failed.json()["detail"]["message"] == "分类服务暂时不可用，请稍后重试"


def test_specific_loan_direction_is_returned_by_classification_endpoint(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case_id = _upload(client).json()["id"]

    def fake_classify(session: Session, case: NationalEconomyClassificationCase):
        result = _loan_result(
            case,
            loan_industry_code="5263",
            loan_industry_name="汽车零配件零售",
            loan_matching_basis=(
                "实际投向用于汽车零部件采购，匹配经营范围内销售汽车零部件，"
                "对应四级代码 5263"
            ),
            loan_matches_enterprise=False,
        )
        session.add(result)
        case.status = "completed"
        session.commit()
        session.refresh(result)
        return result

    monkeypatch.setattr(route_module, "classify_case", fake_classify)

    response = client.post(
        f"/api/v1/national-economy/cases/{case_id}/classifications"
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["loan_industry_code"] == "5263"
    assert payload["loan_industry_name"] == "汽车零配件零售"
    assert "汽车零部件采购" in payload["loan_matching_basis"]
    assert payload["loan_matches_enterprise"] is False


def test_loan_no_match_is_returned_as_needs_review_not_as_enterprise_match(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case_id = _upload(client).json()["id"]
    review_reason = "贷款用途超出营业执照经营范围，贷款投向需人工复核"

    def fake_classify(session: Session, case: NationalEconomyClassificationCase):
        result = _completed_result(case)
        result.status = "needs_review"
        result.loan_matching_basis = review_reason
        session.add(result)
        case.status = "needs_review"
        session.commit()
        session.refresh(result)
        return result

    monkeypatch.setattr(route_module, "classify_case", fake_classify)

    response = client.post(
        f"/api/v1/national-economy/cases/{case_id}/classifications"
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "needs_review"
    assert payload["loan_industry_code"] is None
    assert payload["loan_industry_name"] is None
    assert payload["loan_matching_basis"] == review_reason
    assert payload["loan_matches_enterprise"] is False


def test_matching_loan_direction_is_returned_by_case_and_objection_endpoints(
    client: TestClient,
    db_session: Session,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    case_id = _upload(client).json()["id"]
    case = db_session.get(NationalEconomyClassificationCase, case_id)
    assert case is not None
    db_session.add(
        _loan_result(
            case,
            loan_industry_code="0111",
            loan_industry_name="稻谷种植",
            loan_matching_basis="贷款用途笼统，按企业主营业务判定",
            loan_matches_enterprise=True,
        )
    )
    case.status = "completed"
    db_session.commit()

    case_response = client.get(f"/api/v1/national-economy/cases/{case_id}")

    assert case_response.status_code == 200
    current_result = case_response.json()["current_result"]
    assert current_result["loan_industry_code"] == current_result["industry_code"]
    assert current_result["loan_industry_name"] == current_result["industry_name"]
    assert current_result["loan_matching_basis"] == (
        "贷款用途笼统，按企业主营业务判定"
    )
    assert current_result["loan_matches_enterprise"] is True

    def fake_reclassify(
        session: Session,
        current_case: NationalEconomyClassificationCase,
        objection_text: str,
    ):
        result = _loan_result(
            current_case,
            version=2,
            loan_industry_code="0111",
            loan_industry_name="稻谷种植",
            loan_matching_basis="具体贷款用途命中企业主营业务",
            loan_matches_enterprise=True,
        )
        result.objection = {"description": objection_text}
        session.add(result)
        session.commit()
        session.refresh(result)
        return result

    monkeypatch.setattr(route_module, "reclassify_case", fake_reclassify)
    objection_response = client.post(
        f"/api/v1/national-economy/cases/{case_id}/objections",
        json={"objection_text": "贷款用于主营稻谷种植"},
    )

    assert objection_response.status_code == 200
    objection_payload = objection_response.json()
    assert objection_payload["loan_industry_code"] == "0111"
    assert objection_payload["loan_matches_enterprise"] is True


def test_legacy_result_is_backfilled_by_case_and_history_endpoints(
    client: TestClient,
    db_session: Session,
) -> None:
    case_id = _upload(client).json()["id"]
    case = db_session.get(NationalEconomyClassificationCase, case_id)
    assert case is not None
    legacy_result = _completed_result(case)
    assert legacy_result.loan_industry_code is None
    assert legacy_result.loan_industry_name is None
    assert legacy_result.loan_matching_basis is None
    assert legacy_result.loan_matches_enterprise is None
    db_session.add(legacy_result)
    case.status = "completed"
    db_session.commit()

    case_response = client.get(f"/api/v1/national-economy/cases/{case_id}")
    history_response = client.get(
        f"/api/v1/national-economy/cases/{case_id}/history"
    )

    assert case_response.status_code == 200
    assert history_response.status_code == 200
    expected = {
        "loan_industry_code": "0111",
        "loan_industry_name": "稻谷种植",
        "loan_matching_basis": "贷款投向未单独评估，与企业主营一致",
        "loan_matches_enterprise": True,
    }
    assert {
        key: case_response.json()["current_result"][key] for key in expected
    } == expected
    assert {
        key: history_response.json()["items"][0][key] for key in expected
    } == expected


def test_export_download_contains_three_expected_sheets(
    client: TestClient,
    db_session: Session,
) -> None:
    case_id = _upload(client).json()["id"]
    case = db_session.get(NationalEconomyClassificationCase, case_id)
    assert case is not None
    db_session.add(_completed_result(case))
    case.status = "completed"
    db_session.commit()

    response = client.get(f"/api/v1/national-economy/cases/{case_id}/export")

    assert response.status_code == 200
    assert response.headers["content-type"] == route_module.XLSX_MIME
    assert "attachment" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["案例输入", "当前结论", "判定历史"]

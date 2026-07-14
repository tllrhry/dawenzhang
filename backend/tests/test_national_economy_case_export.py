from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from app.models import (
    AgricultureRelatedResult,
    NationalEconomyClassificationCase,
    NationalEconomyClassificationResult,
)
from app.services.national_economy_case_export import export_case_workbook
from app.services.national_economy_case_ingestion import FIELD_LABELS


def _case() -> NationalEconomyClassificationCase:
    return NationalEconomyClassificationCase(
        id=1,
        scenario="national_economy_classification",
        original_filename="示例企业.docx",
        input_payload={field: f"{label}内容" for field, label in FIELD_LABELS.items()},
        status="needs_review",
    )


def test_export_case_workbook_contains_input_current_result_and_full_history() -> None:
    case = _case()
    NationalEconomyClassificationResult(
        case=case,
        version=3,
        status="needs_review",
        candidate_snapshot=[],
        rationale="候选均不匹配",
        loan_matching_basis="贷款用途超出企业经营范围，需人工复核",
        objection={"description": "主营收入结构已变化"},
    )
    NationalEconomyClassificationResult(
        case=case,
        version=1,
        status="completed",
        industry_code="0111",
        industry_name="稻谷种植",
        confidence=92,
        rationale="主营业务与目录定义一致",
        ai_summary="企业主要从事稻谷种植",
        candidate_snapshot=[],
    )
    NationalEconomyClassificationResult(
        case=case,
        version=2,
        status="completed",
        industry_code="0111",
        industry_major_code="A01",
        industry_name="稻谷种植",
        confidence=94,
        rationale="主营业务与目录定义一致",
        loan_industry_code="0111",
        loan_industry_major_code="A01",
        loan_industry_name="稻谷种植",
        loan_matching_basis="贷款用途笼统，按企业主营业务判定",
        loan_matches_enterprise=True,
        ai_summary="贷款用途与主营一致",
        candidate_snapshot=[],
    )
    NationalEconomyClassificationResult(
        case=case,
        version=4,
        status="completed",
        industry_code="3742",
        industry_major_code="C37",
        industry_name="航天器及运载火箭制造",
        confidence=95,
        rationale="主营业务为航天器制造",
        loan_industry_code="5263",
        loan_industry_major_code="F52",
        loan_industry_name="汽车零配件零售",
        loan_matching_basis="实际投向汽车零部件采购，匹配经营范围内销售汽车零部件",
        loan_matches_enterprise=False,
        ai_summary="贷款投向与企业主营不一致",
        candidate_snapshot=[],
    )

    workbook = load_workbook(BytesIO(export_case_workbook(case)))

    assert workbook.sheetnames == ["案例输入", "当前结论", "判定历史"]

    input_rows = dict(workbook["案例输入"].iter_rows(min_row=2, values_only=True))
    assert input_rows["业务场景"] == "national_economy_classification"
    assert input_rows["原始文件名"] == "示例企业.docx"
    assert input_rows[FIELD_LABELS["enterprise_name"]] == "企业名称内容"
    assert len(input_rows) == len(FIELD_LABELS) + 3

    current_sheet = workbook["当前结论"]
    assert tuple(cell.value for cell in current_sheet[1]) == (
        "行业代码",
        "行业名称",
        "匹配依据",
        "贷款投向代码",
        "贷款投向名称",
        "贷款投向匹配依据",
        "贷款投向是否一致",
    )
    assert tuple(cell.value for cell in current_sheet[2]) == (
        "C37-C3742",
        "航天器及运载火箭制造",
        "主营业务为航天器制造",
        "F52-F5263",
        "汽车零配件零售",
        "实际投向汽车零部件采购，匹配经营范围内销售汽车零部件",
        "不一致",
    )

    history_sheet = workbook["判定历史"]
    assert tuple(cell.value for cell in history_sheet[1]) == (
        "版本号",
        "状态",
        "行业代码",
        "行业名称",
        "匹配依据",
        "贷款投向代码",
        "贷款投向名称",
        "贷款投向匹配依据",
        "贷款投向是否一致",
        "关联异议",
    )
    assert tuple(cell.value for cell in history_sheet[2]) == (
        1,
        "completed",
        "0111",
        "稻谷种植",
        "主营业务与目录定义一致",
        "0111",
        "稻谷种植",
        "贷款投向未单独评估，与企业主营一致",
        "一致",
        None,
    )
    assert tuple(cell.value for cell in history_sheet[3]) == (
        2,
        "completed",
        "A01-A0111",
        "稻谷种植",
        "主营业务与目录定义一致",
        "A01-A0111",
        "稻谷种植",
        "贷款用途笼统，按企业主营业务判定",
        "一致",
        None,
    )
    assert tuple(cell.value for cell in history_sheet[4]) == (
        3,
        "needs_review",
        None,
        None,
        "候选均不匹配",
        None,
        None,
        "贷款用途超出企业经营范围，需人工复核",
        "不一致",
        "主营收入结构已变化",
    )
    assert tuple(cell.value for cell in history_sheet[5]) == (
        4,
        "completed",
        "C37-C3742",
        "航天器及运载火箭制造",
        "主营业务为航天器制造",
        "F52-F5263",
        "汽车零配件零售",
        "实际投向汽车零部件采购，匹配经营范围内销售汽车零部件",
        "不一致",
        None,
    )


def test_export_case_workbook_uses_readable_placeholder_without_completed_result() -> None:
    case = _case()
    NationalEconomyClassificationResult(
        case=case,
        version=1,
        status="needs_review",
        candidate_snapshot=[],
        rationale="需要人工复核",
    )

    workbook = load_workbook(BytesIO(export_case_workbook(case)))

    current_values = tuple(cell.value for cell in workbook["当前结论"][2])
    assert current_values[:2] == (None, None)
    assert current_values[2] == "暂无成功结论（案例状态：needs_review）"
    assert current_values[3:] == (None, None, None, None)


def test_export_case_workbook_reads_back_agriculture_history_and_readable_statuses() -> None:
    case = NationalEconomyClassificationCase(
        id=2,
        scenario="agriculture_related",
        original_filename="涉农示例.docx",
        input_payload={},
        status="completed",
    )
    results = [
        AgricultureRelatedResult(
            id=11,
            case=case,
            scenario_id="agriculture_related",
            version=1,
            status="completed",
            stage_a_result_id=101,
            is_agriculture_related=True,
            matched_categories=[
                {"category": 1, "category_name": "农户贷款", "result": "matched", "method": "rule"},
                {"category": 3, "category_name": "农林牧渔业贷款", "result": "matched", "method": "stage_a"},
                {"category": 2, "category_name": "农村企业及各类组织贷款", "result": "not_matched", "method": "ai"},
            ],
            basis="农户身份字段命中；企业行业门类为农、林、牧、渔业。",
            created_at=datetime(2026, 7, 14, 10, 30),
        ),
        AgricultureRelatedResult(
            id=12,
            case=case,
            scenario_id="agriculture_related",
            version=2,
            status="not_applicable",
            stage_a_result_id=102,
            is_agriculture_related=False,
            matched_categories=[
                {"category": 1, "category_name": "农户贷款", "result": "not_matched", "method": "rule"},
            ],
            basis="四类均未命中。",
        ),
        AgricultureRelatedResult(
            id=13,
            case=case,
            scenario_id="agriculture_related",
            version=3,
            status="needs_review",
            stage_a_result_id=103,
            is_agriculture_related=None,
            matched_categories=[],
            error_detail="注册地址无法判定城乡类别。",
        ),
        AgricultureRelatedResult(
            id=14,
            case=case,
            scenario_id="agriculture_related",
            version=4,
            status="classification_failed",
            stage_a_result_id=104,
            is_agriculture_related=None,
            matched_categories=[],
            error_detail="AI 返回格式不符合约定。",
        ),
    ]

    workbook = load_workbook(
        BytesIO(export_case_workbook(case, agriculture_related_results=results))
    )

    sheet = workbook["涉农判定"]
    assert tuple(cell.value for cell in sheet[1]) == (
        "版本号", "状态", "状态说明", "命中类别", "是否涉农", "匹配依据", "各类别判定方式", "创建时间"
    )
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 4
    assert rows[0][:7] == (
        1, "判定完成", "涉农判定完成。", "农户贷款、农林牧渔业贷款", "是",
        "农户身份字段命中；企业行业门类为农、林、牧、渔业。",
        "类别1（农户贷款）：规则；类别3（农林牧渔业贷款）：Stage A；类别2（农村企业及各类组织贷款）：AI",
    )
    assert rows[1][1:6] == ("不属于涉农", "四类判定均未命中，不属于涉农。", "无", "否", "四类均未命中。")
    assert rows[2][1:6] == ("待人工复核", "涉农判定需人工复核：注册地址无法判定城乡类别。", "无", "未判定", "注册地址无法判定城乡类别。")
    assert rows[3][1:3] == ("判定失败", "涉农判定失败：AI 返回格式不符合约定。")
    assert rows[0][7] == "2026-07-14T10:30:00"

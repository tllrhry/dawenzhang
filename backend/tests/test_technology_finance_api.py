from collections.abc import Iterator
from dataclasses import replace
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.api.routes import national_economy as route_module
from app.core.config import get_settings
from app.db.session import get_db, get_sessionmaker
from app.main import app
from app.models import (
    FiveArticlesResult,
    NationalEconomyClassificationCase,
    NationalEconomyClassificationResult,
)
from app.services.scenario_registry import (
    DIGITAL_FINANCE_REGISTRATION,
    GREEN_FINANCE_REGISTRATION,
    PENSION_FINANCE_REGISTRATION,
    SCENARIO_REGISTRY,
    TECHNOLOGY_FINANCE_FIELD_SCHEMA,
    TECHNOLOGY_FINANCE_REGISTRATION,
    ScenarioRegistration,
)
from app.services.technology_finance_classification_workflow import (
    TechnologyFinanceWorkflowResult,
)


FIXTURES = Path(__file__).parent / "fixtures" / "national_economy"
SCENARIO_ID = "technology_finance"


@pytest.fixture()
def db_session() -> Iterator[Session]:
    session = get_sessionmaker()()
    _clear_cases(session)
    try:
        yield session
    finally:
        session.rollback()
        _clear_cases(session)
        session.close()


@pytest.fixture()
def client(db_session: Session) -> Iterator[TestClient]:
    def override_get_db():
        yield db_session

    app.dependency_overrides[get_db] = override_get_db
    try:
        with TestClient(app) as test_client:
            yield test_client
    finally:
        app.dependency_overrides.pop(get_db, None)


def _clear_cases(session: Session) -> None:
    session.execute(delete(FiveArticlesResult))
    session.execute(delete(NationalEconomyClassificationResult))
    session.execute(delete(NationalEconomyClassificationCase))
    session.commit()


def _upload_technology_finance(client: TestClient):
    template_path = get_settings().technology_finance_template_path
    return client.post(
        f"/api/v1/scenarios/{SCENARIO_ID}/cases",
        files={
            "file": (
                template_path.name,
                template_path.read_bytes(),
                route_module.DOCX_MIME,
            )
        },
    )


def _upload_registered_scenario(
    client: TestClient,
    registration: ScenarioRegistration,
):
    template_path = registration.template_path()
    return client.post(
        f"/api/v1/scenarios/{registration.id}/cases",
        files={
            "file": (
                template_path.name,
                template_path.read_bytes(),
                route_module.DOCX_MIME,
            )
        },
    )


def _upload_national_economy(client: TestClient):
    path = FIXTURES / "valid.docx"
    return client.post(
        "/api/v1/national-economy/cases",
        files={"file": (path.name, path.read_bytes(), route_module.DOCX_MIME)},
    )


def _persist_workflow_result(
    session: Session,
    case: NationalEconomyClassificationCase,
    *,
    version: int,
    objection_text: str | None = None,
) -> TechnologyFinanceWorkflowResult:
    stage_a = NationalEconomyClassificationResult(
        case_id=case.id,
        version=version,
        status="completed",
        industry_code="3011",
        industry_major_code="C30",
        industry_name="水泥制造",
        loan_industry_code="2710",
        loan_industry_major_code="C27",
        loan_industry_name="化学药品原料药制造",
        loan_matching_basis="贷款用于医药项目建设",
        loan_matches_enterprise=False,
        rationale="企业主营命中国民经济行业目录",
        candidate_snapshot=[],
        objection=(
            None if objection_text is None else {"description": objection_text}
        ),
        model_output={"stage": "a"},
    )
    session.add(stage_a)
    session.flush()
    stage_b = FiveArticlesResult(
        case_id=case.id,
        scenario_id=case.scenario,
        version=version,
        status="completed",
        stage_a_result_id=stage_a.id,
        mapping_version_id=None,
        labels=[
            {
                "mapping_version_id": 3,
                "subject": "高技术产业（制造业）",
                "taxonomy_path": ["医药制造业", "化学药品制造"],
                "NEIC_Code": "2710",
                "NEIC_Name": "化学药品原料药制造",
                "source_row": 12,
                "matching_basis": "贷款投向命中科技金融映射。",
                "evidence_refs": [
                    {
                        "type": "mapping",
                        "mapping_version_id": 3,
                        "source_row": 12,
                        "NEIC_Code": "2710",
                        "NEIC_Name": "化学药品原料药制造",
                        "taxonomy_path": ["医药制造业", "化学药品制造"],
                    },
                    {
                        "type": "business",
                        "field_key": "loan_purpose",
                        "field_label": "贷款用途详细描述",
                        "excerpt": "用于医药项目建设",
                    },
                ],
            },
            {
                "mapping_version_id": 3,
                "subject": "战略性新兴产业",
                "taxonomy_path": ["生物产业"],
                "NEIC_Code": "27",
                "NEIC_Name": "医药制造业",
                "source_row": 28,
                "matching_basis": "贷款投向同时命中显式大类映射。",
                "evidence_refs": [
                    {
                        "type": "mapping",
                        "mapping_version_id": 3,
                        "source_row": 28,
                        "NEIC_Code": "27",
                        "NEIC_Name": "医药制造业",
                        "taxonomy_path": ["生物产业"],
                    },
                    {
                        "type": "business",
                        "field_key": "stage_a.loan_matching_basis",
                        "field_label": "Stage A 贷款投向匹配依据",
                        "excerpt": "贷款用于医药项目建设",
                    },
                ],
            },
        ],
        loan_neic_code="2710",
        loan_neic_name="化学药品原料药制造",
        enterprise_neic_code="3011",
        enterprise_neic_name="水泥制造",
        consistency_status="consistent",
        consistency_basis="贷款用途服务于企业科技活动。",
        consistency_evidence_refs=[],
        model_output={"stage": "b"},
    )
    session.add(stage_b)
    case.status = "completed"
    session.commit()
    session.refresh(stage_a)
    session.refresh(stage_b)
    return TechnologyFinanceWorkflowResult(stage_a, stage_b)


def test_technology_finance_seven_endpoint_types(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    template_response = client.get(f"/api/v1/scenarios/{SCENARIO_ID}/template")
    assert template_response.status_code == 200
    assert template_response.content == (
        get_settings().technology_finance_template_path.read_bytes()
    )

    upload_response = _upload_technology_finance(client)
    assert upload_response.status_code == 201
    case_id = upload_response.json()["id"]

    detail_response = client.get(f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}")
    assert detail_response.status_code == 200
    assert [item["field"] for item in detail_response.json()["input_fields"]] == [
        field.key for field in TECHNOLOGY_FINANCE_FIELD_SCHEMA
    ]

    def fake_classify(session: Session, case: NationalEconomyClassificationCase):
        return _persist_workflow_result(session, case, version=1)

    monkeypatch.setattr(
        route_module,
        "classify_technology_finance_case",
        fake_classify,
    )
    classification_response = client.post(
        f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}/classifications"
    )
    assert classification_response.status_code == 200
    classification = classification_response.json()
    assert classification["stage_a"]["status"] == "completed"
    assert classification["stage_b"]["status"] == "completed"
    assert classification["stage_b"]["labels"][0]["source_row"] == 12
    assert classification["stage_b"]["consistency_status"] == "consistent"

    def fake_reclassify(
        session: Session,
        case: NationalEconomyClassificationCase,
        objection_text: str,
    ):
        return _persist_workflow_result(
            session,
            case,
            version=2,
            objection_text=objection_text,
        )

    monkeypatch.setattr(
        route_module,
        "reclassify_technology_finance_case",
        fake_reclassify,
    )
    objection_response = client.post(
        f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}/objections",
        json={"objection_text": "请按医药项目重新判定"},
    )
    assert objection_response.status_code == 200
    assert objection_response.json()["stage_a"]["version"] == 2
    assert objection_response.json()["stage_b"]["stage_a_result_id"] == (
        objection_response.json()["stage_a"]["id"]
    )

    history_response = client.get(
        f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}/history"
    )
    assert history_response.status_code == 200
    history = history_response.json()["items"]
    assert [item["version"] for item in history] == [1, 2]
    assert all(item["stage_a_result_id"] for item in history)

    export_response = client.get(
        f"/api/v1/scenarios/{SCENARIO_ID}/cases/{case_id}/export"
    )
    assert export_response.status_code == 200
    assert export_response.headers["content-type"] == route_module.XLSX_MIME
    workbook = load_workbook(BytesIO(export_response.content))
    assert workbook.sheetnames == [
        "案例输入",
        "当前结论",
        "判定历史",
        "科技金融判定",
    ]
    exported_labels = {
        cell.value for cell in workbook["案例输入"]["A"] if cell.value is not None
    }
    assert {field.label for field in TECHNOLOGY_FINANCE_FIELD_SCHEMA} <= exported_labels
    technology_sheet = workbook["科技金融判定"]
    headers = tuple(cell.value for cell in technology_sheet[1])
    rows = [dict(zip(headers, row, strict=True)) for row in technology_sheet.iter_rows(
        min_row=2,
        values_only=True,
    )]
    assert [row["主题"] for row in rows] == [
        "高技术产业（制造业）",
        "战略性新兴产业",
    ]
    assert [row["映射源行"] for row in rows] == [12, 28]
    assert rows[0]["业务证据摘要"] == "贷款用途详细描述：用于医药项目建设"
    assert rows[1]["Stage A结果ID"] == objection_response.json()["stage_a"]["id"]


@pytest.mark.parametrize(
    "registration",
    [
        TECHNOLOGY_FINANCE_REGISTRATION,
        GREEN_FINANCE_REGISTRATION,
        DIGITAL_FINANCE_REGISTRATION,
        PENSION_FINANCE_REGISTRATION,
    ],
    ids=lambda registration: registration.id,
)
def test_registered_scenario_upload_and_detail_return_complete_profile_schema(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
    registration: ScenarioRegistration,
) -> None:
    available_registration = replace(registration, status="available")
    monkeypatch.setattr(
        route_module,
        "SCENARIO_REGISTRY",
        {
            **SCENARIO_REGISTRY,
            registration.id: available_registration,
        },
    )

    upload_response = _upload_registered_scenario(client, available_registration)

    assert upload_response.status_code == 201
    created = upload_response.json()
    assert created["scenario"] == registration.id
    detail_response = client.get(
        f"/api/v1/scenarios/{registration.id}/cases/{created['id']}"
    )
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert detail["scenario"] == registration.id
    assert [(item["field"], item["label"]) for item in detail["input_fields"]] == [
        (field.key, field.label) for field in registration.field_schema
    ]
    assert len(detail["input_fields"]) == len(registration.field_schema)


@pytest.mark.parametrize("scenario_id", ["inclusive_finance", "not_registered"])
def test_unavailable_or_unknown_upload_is_rejected_before_handler_dispatch(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
    scenario_id: str,
) -> None:
    handler_lookup = MagicMock()
    monkeypatch.setattr(route_module, "get_scenario_case_handler", handler_lookup)
    template_path = get_settings().technology_finance_template_path

    response = client.post(
        f"/api/v1/scenarios/{scenario_id}/cases",
        files={
            "file": (
                template_path.name,
                template_path.read_bytes(),
                route_module.DOCX_MIME,
            )
        },
    )

    assert response.status_code == (409 if scenario_id == "inclusive_finance" else 404)
    handler_lookup.assert_not_called()


def test_scenario_mismatch_is_rejected_before_detail_handler_dispatch(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    national_case_id = _upload_national_economy(client).json()["id"]
    handler_lookup = MagicMock()
    monkeypatch.setattr(route_module, "get_scenario_case_handler", handler_lookup)

    response = client.get(
        f"/api/v1/scenarios/{SCENARIO_ID}/cases/{national_case_id}"
    )

    assert response.status_code == 404
    assert response.json()["detail"] == "案例不存在"
    handler_lookup.assert_not_called()


@pytest.mark.parametrize(
    "scenario_id",
    [
        "agriculture_related",
        "green_finance",
        "inclusive_finance",
        "pension_finance",
        "digital_finance",
    ],
)
def test_coming_soon_scenarios_are_rejected(
    client: TestClient,
    scenario_id: str,
) -> None:
    response = client.get(f"/api/v1/scenarios/{scenario_id}/template")

    assert response.status_code == 409
    assert response.json()["detail"] == "场景暂未开放"


def test_coming_soon_case_upload_is_rejected(client: TestClient) -> None:
    template_path = get_settings().technology_finance_template_path

    response = client.post(
        "/api/v1/scenarios/green_finance/cases",
        files={
            "file": (
                template_path.name,
                template_path.read_bytes(),
                route_module.DOCX_MIME,
            )
        },
    )

    assert response.status_code == 409
    assert response.json()["detail"] == "场景暂未开放"


@pytest.mark.parametrize(
    ("method", "suffix", "json"),
    [
        ("GET", "", None),
        ("POST", "/classifications", None),
        ("POST", "/objections", {"objection_text": "场景错配"}),
        ("GET", "/history", None),
        ("GET", "/export", None),
    ],
)
def test_scenario_case_mismatch_is_rejected_before_workflow(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
    method: str,
    suffix: str,
    json: dict[str, str] | None,
) -> None:
    national_case_id = _upload_national_economy(client).json()["id"]
    classifier = MagicMock()
    reclassifier = MagicMock()
    monkeypatch.setattr(route_module, "classify_technology_finance_case", classifier)
    monkeypatch.setattr(
        route_module,
        "reclassify_technology_finance_case",
        reclassifier,
    )

    response = client.request(
        method,
        f"/api/v1/scenarios/{SCENARIO_ID}/cases/{national_case_id}{suffix}",
        json=json,
    )

    assert response.status_code == 404
    assert response.json()["detail"] == "案例不存在"
    classifier.assert_not_called()
    reclassifier.assert_not_called()


def test_unknown_scenario_is_not_treated_as_coming_soon(client: TestClient) -> None:
    response = client.get("/api/v1/scenarios/not_registered/template")

    assert response.status_code == 404
    assert response.json()["detail"] == "场景不存在"

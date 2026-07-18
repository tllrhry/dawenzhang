from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import (
    FiveArticlesResult,
    NationalEconomyClassificationCase,
    NationalEconomyClassificationResult,
)
from app.services.national_economy_case_export import export_case_workbook
from app.services.scenario_registry import (
    DIGITAL_FINANCE_REGISTRATION,
    GREEN_FINANCE_REGISTRATION,
    PENSION_FINANCE_REGISTRATION,
    TECHNOLOGY_FINANCE_FIELD_SCHEMA,
    TECHNOLOGY_FINANCE_SCENARIO,
    ScenarioRegistration,
)


def _case_with_stage_a(
    *,
    stage_a_id: int = 11,
    stage_a_version: int = 1,
) -> tuple[NationalEconomyClassificationCase, NationalEconomyClassificationResult]:
    case = NationalEconomyClassificationCase(
        id=1,
        scenario=TECHNOLOGY_FINANCE_SCENARIO,
        original_filename="科技金融案例.docx",
        input_payload={
            field.key: f"{field.label}内容" for field in TECHNOLOGY_FINANCE_FIELD_SCHEMA
        },
        status="completed",
    )
    stage_a = NationalEconomyClassificationResult(
        id=stage_a_id,
        case=case,
        version=stage_a_version,
        status="completed",
        industry_code="3011",
        industry_major_code="C30",
        industry_name="水泥制造",
        loan_industry_code="2710",
        loan_industry_major_code="C27",
        loan_industry_name="化学药品原料药制造",
        loan_matching_basis="贷款用于医药项目建设",
        loan_matches_enterprise=False,
        rationale="企业主营命中国民经济行业目录",
        candidate_snapshot=[],
    )
    return case, stage_a


def _completed_stage_b(
    *,
    stage_a_result_id: int,
    consistency_status: str,
) -> FiveArticlesResult:
    return FiveArticlesResult(
        id=21,
        case_id=1,
        scenario_id=TECHNOLOGY_FINANCE_SCENARIO,
        version=1,
        status="completed",
        stage_a_result_id=stage_a_result_id,
        mapping_version_id=3,
        labels=[
            {
                "mapping_version_id": 3,
                "subject": "高技术产业（制造业）",
                "taxonomy_path": [
                    "医药制造业",
                    "化学药品制造",
                    "原料药制造",
                    "创新药原料",
                ],
                "NEIC_Code": "2710",
                "NEIC_Name": "化学药品原料药制造",
                "source_row": 12,
                "matching_basis": "贷款用于创新药原料项目，命中科技金融映射。",
                "evidence_refs": [
                    {
                        "type": "mapping",
                        "mapping_version_id": 3,
                        "source_row": 12,
                        "NEIC_Code": "2710",
                        "NEIC_Name": "化学药品原料药制造",
                        "taxonomy_path": [
                            "医药制造业",
                            "化学药品制造",
                            "原料药制造",
                            "创新药原料",
                        ],
                    },
                    {
                        "type": "business",
                        "field_key": "loan_purpose",
                        "field_label": "贷款用途详细描述",
                        "excerpt": "用于创新药原料项目建设",
                    },
                    {
                        "type": "business",
                        "field_key": "qualification_certification",
                        "field_label": "企业核心资质与认证",
                        "excerpt": "高新技术企业认证",
                    },
                ],
            },
            {
                "mapping_version_id": 3,
                "subject": "战略性新兴产业",
                "taxonomy_path": ["生物产业", "生物医药产业"],
                "NEIC_Code": "27",
                "NEIC_Name": "医药制造业",
                "source_row": 28,
                "matching_basis": "贷款投向命中显式医药制造业大类映射。",
                "evidence_refs": [
                    {
                        "type": "mapping",
                        "mapping_version_id": 3,
                        "source_row": 28,
                        "NEIC_Code": "27",
                        "NEIC_Name": "医药制造业",
                        "taxonomy_path": ["生物产业", "生物医药产业"],
                    },
                    {
                        "type": "business",
                        "field_key": "stage_a.loan_matching_basis",
                        "field_label": "Stage A 贷款投向匹配依据",
                        "excerpt": "贷款用于医药项目建设",
                    },
                ],
            },
        ],
        loan_neic_code="2710",
        loan_neic_name="化学药品原料药制造",
        enterprise_neic_code="3011",
        enterprise_neic_name="水泥制造",
        consistency_status=consistency_status,
        consistency_basis="已综合企业类别、贷款用途与 Stage A 投向证据判断。",
        consistency_evidence_refs=[],
    )


@pytest.mark.parametrize(
    ("consistency_status", "display_name"),
    [
        ("consistent", "一致"),
        ("inconsistent", "不一致"),
        ("needs_review", "待人工复核"),
    ],
)
def test_completed_export_reads_back_multi_labels_sources_evidence_and_consistency(
    consistency_status: str,
    display_name: str,
) -> None:
    case, stage_a = _case_with_stage_a()
    stage_b = _completed_stage_b(
        stage_a_result_id=stage_a.id,
        consistency_status=consistency_status,
    )

    workbook = load_workbook(
        BytesIO(
            export_case_workbook(
                case,
                five_articles_results=[stage_b],
            )
        )
    )

    assert workbook.sheetnames == [
        "案例输入",
        "当前结论",
        "判定历史",
        "科技金融判定",
    ]
    input_rows = dict(workbook["案例输入"].iter_rows(min_row=2, values_only=True))
    assert {field.label for field in TECHNOLOGY_FINANCE_FIELD_SCHEMA} <= set(input_rows)

    sheet = workbook["科技金融判定"]
    headers = tuple(cell.value for cell in sheet[1])
    rows = [
        dict(zip(headers, row, strict=True))
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]
    assert len(rows) == 2
    assert rows[0]["主题"] == "高技术产业（制造业）"
    assert [rows[0][f"第{name}层"] for name in ("一", "二", "三", "四")] == [
        "医药制造业",
        "化学药品制造",
        "原料药制造",
        "创新药原料",
    ]
    assert rows[1]["第三层"] is None
    assert [row["映射源行"] for row in rows] == [12, 28]
    assert rows[0]["映射代码"] == "2710"
    assert rows[0]["映射名称"] == "化学药品原料药制造"
    assert rows[0]["业务证据摘要"] == (
        "贷款用途详细描述：用于创新药原料项目建设\n"
        "企业核心资质与认证：高新技术企业认证"
    )
    assert "mapping" not in rows[0]["业务证据摘要"]
    assert rows[0]["Stage A结果ID"] == stage_a.id
    assert all(
        row["贷款对应的五篇大文章类别与企业类别是否一致"] == display_name
        for row in rows
    )
    assert all(row["一致性依据"] == stage_b.consistency_basis for row in rows)


def test_export_writes_ip_intensive_industry_condition_per_label() -> None:
    case, stage_a = _case_with_stage_a()
    stage_b = _completed_stage_b(
        stage_a_result_id=stage_a.id,
        consistency_status="consistent",
    )
    stage_b.labels[0] = {
        **stage_b.labels[0],
        "subject": "知识产权(专利)密集型产业",
        "ip_intensive_industry_status": "satisfied",
        "ip_intensive_industry_basis": "企业名称已在名录中匹配到。",
    }
    stage_b.labels[1] = {
        **stage_b.labels[1],
        "subject": "知识产权(专利)密集型产业",
        "ip_intensive_industry_status": "unsatisfied",
        "ip_intensive_industry_basis": "企业名称未在名录中匹配到。",
    }

    workbook = load_workbook(
        BytesIO(export_case_workbook(case, five_articles_results=[stage_b]))
    )
    sheet = workbook["科技金融判定"]
    headers = tuple(cell.value for cell in sheet[1])
    rows = [
        dict(zip(headers, row, strict=True))
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

    assert rows[0]["知识产权条件"] == "满足"
    assert rows[1]["知识产权条件"] == "不满足：企业名称未在名录中匹配到。"


def test_technology_export_writes_direction_and_auxiliary_evidence_columns() -> None:
    case, stage_a = _case_with_stage_a()
    stage_b = _completed_stage_b(
        stage_a_result_id=stage_a.id,
        consistency_status="consistent",
    )
    stage_b.consistency_evidence_refs = [
        {
            "type": "technology_direction",
            "mapping_hit": True,
            "NEIC_Code": "2710",
            "NEIC_Name": "化学药品原料药制造",
            "taxonomy_path": ["高技术产业", "医药制造"],
        },
        {
            "type": "technology_auxiliary",
            "evidence_role": "official_qualification",
            "status": "satisfied",
            "excerpt": "高新技术企业",
        },
        {
            "type": "technology_registry",
            "registry_type": "high_tech",
            "status": "satisfied",
            "matched": True,
            "excerpt": "企业名称已在高新技术企业名单中匹配到（来源序号 1）。",
        },
        {
            "type": "technology_registry",
            "registry_type": "specialized_innovation",
            "status": "unsatisfied",
            "matched": False,
            "excerpt": "企业名称未在专精特新企业名单中匹配到。",
        },
        {
            "type": "technology_auxiliary",
            "evidence_role": "rd_staff_ratio",
            "status": "satisfied",
            "normalized_percent": 10.0,
        },
        {
            "type": "technology_auxiliary",
            "evidence_role": "rd_investment_ratio",
            "status": "unsatisfied",
            "normalized_amount_wan": 299.0,
            "derived_ratio_percent": 2.99,
            "warning": "研发投入占营收比例低于3%参考阈值",
        },
        {
            "type": "technology_auxiliary",
            "evidence_role": "patent_software_copyright",
            "status": "satisfied",
            "excerpt": "拥有发明专利3项",
        },
    ]

    workbook = load_workbook(
        BytesIO(export_case_workbook(case, five_articles_results=[stage_b]))
    )
    sheet = workbook["科技金融判定"]
    headers = tuple(cell.value for cell in sheet[1])
    row = dict(zip(headers, tuple(cell.value for cell in sheet[2]), strict=True))

    assert row["贷款实际投向依据"] == (
        "命中科技金融映射：2710 化学药品原料药制造 · 高技术产业 / 医药制造"
    )
    assert row["官方科技资质"] == "满足：高新技术企业"
    assert row["高新技术企业名单"] == (
        "满足：企业名称已在高新技术企业名单中匹配到（来源序号 1）。"
    )
    assert row["专精特新企业名单"] == (
        "未满足：企业名称未在专精特新企业名单中匹配到。"
    )
    assert row["研发人员占比"] == "满足：10%"
    assert row["研发投入"] == "299万元"
    assert row["研发投入占营收比例"] == "未满足：2.99%"
    assert row["专利或软著等"] == "满足：拥有发明专利3项"
    assert row["辅助证据预警"] == "研发投入占营收比例低于3%参考阈值"


@pytest.mark.parametrize(
    (
        "result_status",
        "consistency_status",
        "status_label",
        "consistency_label",
        "error_detail",
    ),
    [
        (
            "not_applicable",
            "not_applicable",
            "不属于科技金融",
            "不适用",
            "loan_direction_has_no_explicit_mapping",
        ),
        (
            "needs_review",
            "needs_review",
            "待人工复核",
            "待人工复核",
            "published_mapping_version_not_found",
        ),
        (
            "classification_failed",
            None,
            "判定失败",
            "不适用",
            "DeepSeek request timed out",
        ),
    ],
)
def test_export_without_formal_labels_has_readable_status_and_stage_a_context(
    result_status: str,
    consistency_status: str | None,
    status_label: str,
    consistency_label: str,
    error_detail: str,
) -> None:
    case, stage_a = _case_with_stage_a()
    stage_b = FiveArticlesResult(
        id=22,
        case_id=case.id,
        scenario_id=case.scenario,
        version=2,
        status=result_status,
        stage_a_result_id=stage_a.id,
        mapping_version_id=3,
        labels=[],
        loan_neic_code="2710",
        loan_neic_name="化学药品原料药制造",
        enterprise_neic_code="3011",
        enterprise_neic_name="水泥制造",
        consistency_status=consistency_status,
        consistency_basis=(
            "贷款投向未命中映射，一致性不适用。"
            if result_status == "not_applicable"
            else "科技金融映射数据异常，需人工复核。"
        ),
        consistency_evidence_refs=[],
        error_detail=error_detail,
    )

    workbook = load_workbook(
        BytesIO(export_case_workbook(case, five_articles_results=[stage_b]))
    )
    sheet = workbook["科技金融判定"]
    headers = tuple(cell.value for cell in sheet[1])
    row = dict(zip(headers, tuple(cell.value for cell in sheet[2]), strict=True))

    assert row["科技金融状态"] == status_label
    assert row["状态说明"]
    assert row["主题"] is None
    assert row["映射源行"] is None
    assert row["Stage A结果ID"] == stage_a.id
    assert row["贷款投向国民经济行业代码"] == "2710"
    assert row["企业国民经济行业名称"] == "水泥制造"
    assert row["贷款对应的五篇大文章类别与企业类别是否一致"] == (
        consistency_label
    )
    assert "不适用" in row["一致性依据"]


def test_export_uses_latest_stage_b_version_and_its_stage_a_association() -> None:
    case, first_stage_a = _case_with_stage_a()
    second_stage_a = NationalEconomyClassificationResult(
        id=12,
        case=case,
        version=2,
        status="completed",
        industry_code="2710",
        industry_name="化学药品原料药制造",
        loan_industry_code="2720",
        loan_industry_name="化学药品制剂制造",
        loan_matching_basis="异议后按制剂项目判定",
        rationale="异议重判后的 Stage A 依据",
        candidate_snapshot=[],
    )
    older = _completed_stage_b(
        stage_a_result_id=first_stage_a.id,
        consistency_status="consistent",
    )
    latest = FiveArticlesResult(
        id=23,
        case_id=case.id,
        scenario_id=case.scenario,
        version=2,
        status="classification_failed",
        stage_a_result_id=second_stage_a.id,
        labels=[],
        loan_neic_code="2720",
        loan_neic_name="化学药品制剂制造",
        enterprise_neic_code="2710",
        enterprise_neic_name="化学药品原料药制造",
        consistency_evidence_refs=[],
        error_detail="模型输出校验失败",
    )

    workbook = load_workbook(
        BytesIO(
            export_case_workbook(
                case,
                five_articles_results=[latest, older],
            )
        )
    )
    sheet = workbook["科技金融判定"]
    headers = tuple(cell.value for cell in sheet[1])
    row = dict(zip(headers, tuple(cell.value for cell in sheet[2]), strict=True))

    assert row["Stage B版本"] == 2
    assert row["Stage A结果ID"] == second_stage_a.id
    assert row["贷款投向国民经济行业代码"] == "2720"
    assert row["科技金融状态"] == "判定失败"


@pytest.mark.parametrize(
    "profile",
    [
        GREEN_FINANCE_REGISTRATION,
        DIGITAL_FINANCE_REGISTRATION,
        PENSION_FINANCE_REGISTRATION,
    ],
    ids=lambda profile: profile.id,
)
@pytest.mark.parametrize(
    ("result_status", "status_label"),
    [
        ("completed", "判定完成"),
        ("not_applicable", None),
        ("needs_review", "待人工复核"),
        ("classification_failed", "判定失败"),
    ],
)
def test_new_finance_export_uses_profile_sheet_and_readable_statuses(
    profile: ScenarioRegistration,
    result_status: str,
    status_label: str | None,
) -> None:
    case = NationalEconomyClassificationCase(
        id=1,
        scenario=profile.id,
        original_filename=f"{profile.name}案例.docx",
        input_payload={
            field.key: f"{field.label}内容" for field in profile.field_schema
        },
        status="completed",
    )
    stage_a = NationalEconomyClassificationResult(
        id=11,
        case=case,
        version=1,
        status="completed",
        industry_code="3011",
        industry_name="水泥制造",
        loan_industry_code="2710",
        loan_industry_name="化学药品原料药制造",
        loan_matching_basis="贷款用于项目建设",
        rationale="企业主营命中国民经济行业目录",
        candidate_snapshot=[],
    )
    labels = (
        [
            {
                "subject": f"{profile.name}主题",
                "taxonomy_path": ["第一层", "第二层"],
                "NEIC_Code": "2710",
                "NEIC_Name": "化学药品原料药制造",
                "source_row": 12,
                "matching_basis": f"命中{profile.name}映射。",
                "evidence_refs": [
                    {
                        "type": "business",
                        "field_key": "loan_purpose",
                        "field_label": "贷款用途详细描述",
                        "excerpt": "用于项目建设",
                    }
                ],
            }
        ]
        if result_status == "completed"
        else []
    )
    stage_b = FiveArticlesResult(
        id=21,
        case_id=case.id,
        scenario_id=profile.id,
        version=1,
        status=result_status,
        stage_a_result_id=stage_a.id,
        mapping_version_id=3,
        labels=labels,
        loan_neic_code="2710",
        loan_neic_name="化学药品原料药制造",
        enterprise_neic_code="3011",
        enterprise_neic_name="水泥制造",
        consistency_status=(
            "consistent" if result_status == "completed" else "not_applicable"
        ),
        consistency_basis="已综合场景业务证据判断。",
        consistency_evidence_refs=[],
        error_detail="需要人工确认" if result_status == "needs_review" else None,
    )

    workbook = load_workbook(
        BytesIO(
            export_case_workbook(
                case,
                five_articles_results=[stage_b],
                profile=profile,
            )
        )
    )

    assert profile.export_sheet_name in workbook.sheetnames
    assert {"案例输入", "当前结论", "判定历史"} <= set(workbook.sheetnames)
    sheet = workbook[profile.export_sheet_name]
    headers = tuple(cell.value for cell in sheet[1])
    row = dict(zip(headers, tuple(cell.value for cell in sheet[2]), strict=True))
    assert row[f"{profile.name}状态"] == (
        status_label or f"不属于{profile.name}"
    )
    assert row["Stage A结果ID"] == stage_a.id
    assert profile.name in row["状态说明"]
    if result_status == "completed":
        assert row["映射源行"] == 12
        assert row["业务证据摘要"] == "贷款用途详细描述：用于项目建设"
        assert row["一致性依据"] == stage_b.consistency_basis
    else:
        assert row["映射源行"] is None
        assert row["业务证据摘要"] is None
        assert row["一致性依据"]


def test_digital_export_includes_category_policy_auxiliary_raw_text_and_warnings() -> None:
    profile = DIGITAL_FINANCE_REGISTRATION
    case = NationalEconomyClassificationCase(
        id=81,
        scenario=profile.id,
        original_filename="数字金融案例.docx",
        input_payload={field.key: "" for field in profile.field_schema},
        status="completed",
    )
    result = FiveArticlesResult(
        id=101,
        case_id=case.id,
        scenario_id=profile.id,
        version=1,
        status="completed",
        stage_a_result_id=91,
        mapping_version_id=7,
        decision_policy_version="digital-direction-v1",
        labels=[
            {
                "subject": "数字经济及其核心产业",
                "taxonomy_path": ["数字化效率提升业", "智能制造"],
                "NEIC_Code": "6513",
                "NEIC_Name": "应用软件开发",
                "source_row": 12,
                "digital_category": "产业数字化",
                "matching_basis": "贷款实际投向命中产业数字化服务器归一规则。",
                "evidence_refs": [],
            }
        ],
        loan_neic_code="6513",
        loan_neic_name="应用软件开发",
        enterprise_neic_code="3011",
        enterprise_neic_name="水泥制造",
        consistency_status="inconsistent",
        consistency_basis="以该笔资金投向为准。",
        consistency_evidence_refs=[
            {
                "type": "digital_direction",
                "digital_category": "产业数字化",
            },
            {
                "type": "digital_auxiliary",
                "evidence_role": "industry_positioning",
                "excerpt": "传统制造企业数字化转型主体",
                "warning": None,
            },
            {
                "type": "digital_auxiliary",
                "evidence_role": "core_competitiveness",
                "excerpt": "仅采购外部平台",
                "warning": "企业数字化核心竞争力佐证不足",
            },
            {
                "type": "digital_auxiliary",
                "evidence_role": "rd_ip",
                "excerpt": "仅有商标",
                "warning": "知识产权佐证不足",
            },
        ],
        model_output={"digital_decision": {"digital_category": "产业数字化"}},
    )

    workbook = load_workbook(
        BytesIO(
            export_case_workbook(
                case,
                five_articles_results=[result],
                profile=profile,
            )
        )
    )
    sheet = workbook[profile.export_sheet_name]
    headers = tuple(cell.value for cell in sheet[1])
    row = dict(zip(headers, tuple(cell.value for cell in sheet[2]), strict=True))

    assert row["数字类别"] == "产业数字化"
    assert row["数字决策策略版本"] == "digital-direction-v1"
    assert row["行业定位原文"] == "传统制造企业数字化转型主体"
    assert row["数字核心竞争力原文"] == "仅采购外部平台"
    assert row["研发知识产权原文"] == "仅有商标"
    assert row["辅助证据预警"] == (
        "企业数字化核心竞争力佐证不足；知识产权佐证不足"
    )


def test_green_export_includes_directory_match_method_auxiliary_and_violation() -> None:
    profile = GREEN_FINANCE_REGISTRATION
    case = NationalEconomyClassificationCase(
        id=82,
        scenario=profile.id,
        original_filename="绿色金融案例.docx",
        input_payload={field.key: "" for field in profile.field_schema},
        status="completed",
    )
    result = FiveArticlesResult(
        id=102,
        case_id=case.id,
        scenario_id=profile.id,
        version=1,
        status="needs_review",
        stage_a_result_id=92,
        mapping_version_id=7,
        decision_policy_version="green-direction-v3",
        labels=[
            {
                "subject": "绿色金融支持项目目录（2025年版）",
                "taxonomy_path": ["清洁能源产业", "太阳能利用"],
                "NEIC_Code": "4415",
                "NEIC_Name": "太阳能发电",
                "source_row": 12,
                "match_method": "condition_fallback",
                "matching_basis": "贷款实际投向命中绿色目录。",
                "evidence_refs": [],
            }
        ],
        loan_neic_code="4415",
        loan_neic_name="太阳能发电",
        enterprise_neic_code="3011",
        enterprise_neic_name="水泥制造",
        consistency_status="needs_review",
        consistency_basis="存在重大环保违法失信，转人工尽调。",
        consistency_evidence_refs=[
            {
                "type": "green_direction",
                "subject": "绿色金融支持项目目录（2025年版）",
                "taxonomy_path": ["清洁能源产业", "太阳能利用"],
                "match_method": "condition_fallback",
            },
            {
                "type": "green_auxiliary",
                "field_key": "green_certifications",
                "excerpt": "绿色项目认定文件",
                "warning": None,
            },
            {
                "type": "green_auxiliary",
                "field_key": "energy_saving_pollution_control",
                "excerpt": "",
                "warning": "缺少节能减排/污染治理内容",
            },
            {
                "type": "green_auxiliary",
                "field_key": "carbon_environmental_benefits",
                "excerpt": "预计年减排1200吨",
                "warning": None,
            },
            {
                "type": "green_violation",
                "raw_value": "有",
                "violation_status": "yes",
                "warning": "存在重大环保违法失信，需开展人工尽职调查并核验潜在漂绿风险",
            },
        ],
    )

    workbook = load_workbook(
        BytesIO(
            export_case_workbook(
                case,
                five_articles_results=[result],
                profile=profile,
            )
        )
    )
    sheet = workbook[profile.export_sheet_name]
    headers = tuple(cell.value for cell in sheet[1])
    row = dict(zip(headers, tuple(cell.value for cell in sheet[2]), strict=True))

    assert row["绿色目录标签"] == (
        "绿色金融支持项目目录（2025年版） / 清洁能源产业 / 太阳能利用"
    )
    assert row["条件匹配方式"] == "条件回退命中"
    assert row["绿色决策策略版本"] == "green-direction-v3"
    assert row["环保与绿色资质认证原文"] == "绿色项目认定文件"
    assert row["节能减排污染治理原文"] is None
    assert row["碳排放与环境效益原文"] == "预计年减排1200吨"
    assert row["重大环保违法失信原文"] == "有"
    assert row["重大环保违法失信状态"] == "yes"
    assert row["辅助证据预警"] == (
        "缺少节能减排/污染治理内容；"
        "存在重大环保违法失信，需开展人工尽职调查并核验潜在漂绿风险"
    )


def test_pension_export_includes_matrix_shares_subject_basis_and_warning() -> None:
    profile = PENSION_FINANCE_REGISTRATION
    case = NationalEconomyClassificationCase(
        id=1,
        scenario=profile.id,
        original_filename="养老金融案例.docx",
        input_payload={
            field.key: f"{field.label}内容" for field in profile.field_schema
        },
        status="completed",
    )
    result = FiveArticlesResult(
        id=21,
        case_id=case.id,
        scenario_id=profile.id,
        version=1,
        status="completed",
        stage_a_result_id=11,
        mapping_version_id=3,
        decision_policy_version="pension-direction-share-v2",
        labels=[
            {
                "subject": "养老产业",
                "taxonomy_path": ["养老服务"],
                "NEIC_Code": "8514",
                "NEIC_Name": "老年人、残疾人养护服务",
                "source_row": 12,
                "matching_basis": "养老投向矩阵认定。",
                "evidence_refs": [],
            }
        ],
        loan_neic_code="8514",
        loan_neic_name="老年人、残疾人养护服务",
        enterprise_neic_code="7020",
        enterprise_neic_name="物业管理",
        consistency_status="inconsistent",
        consistency_basis="养老营收占比达到50%，以主体属性辅助认定。",
        consistency_evidence_refs=[
            {
                "type": "pension_matrix",
                "field_key": "pension_loan_direction_share",
                "raw_value": "",
                "normalized_percent": None,
                "matrix_branch": "PENSION_REVENUE_AT_LEAST_50_UNKNOWN_LOAN_SHARE",
            },
            {
                "type": "pension_matrix",
                "field_key": "main_business_revenue_share",
                "raw_value": "养老服务占60%",
                "normalized_percent": 60.0,
                "matrix_branch": "PENSION_REVENUE_AT_LEAST_50_UNKNOWN_LOAN_SHARE",
            },
            {
                "type": "pension_qualification",
                "warning": "未提供养老许可、备案或重点项目清单等辅助资质",
            },
        ],
    )

    workbook = load_workbook(
        BytesIO(
            export_case_workbook(
                case,
                five_articles_results=[result],
                profile=profile,
            )
        )
    )
    sheet = workbook[profile.export_sheet_name]
    headers = tuple(cell.value for cell in sheet[1])
    row = dict(zip(headers, tuple(cell.value for cell in sheet[2]), strict=True))

    assert row["养老矩阵分支"] == "PENSION_REVENUE_AT_LEAST_50_UNKNOWN_LOAN_SHARE"
    assert row["贷款养老投向占比规范化"] is None
    assert row["主营业务及营收占比原始值"] == "养老服务占60%"
    assert row["主营业务及营收占比规范化"] == "60.0%"
    assert row["主体辅助依据"] == "养老产业营业收入占比达到50%（含）"
    assert row["养老资质预警"] == "未提供养老许可、备案或重点项目清单等辅助资质"

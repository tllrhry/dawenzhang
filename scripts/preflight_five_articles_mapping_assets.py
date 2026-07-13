#!/usr/bin/env python3
"""Validate the local green, digital, and pension finance mapping assets."""

from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
import re
import sys
import unicodedata
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


ROOT = Path(__file__).resolve().parents[1]
ASSET_DIR = ROOT / "五篇大文章映射"

REQUIRED_HEADERS = (
    "主题",
    "第一层",
    "第二层",
    "第三层",
    "第四层",
    "NEIC_Code",
    "NEIC_Name",
)
OPTIONAL_CATEGORY_HEADER = "属于类别"
ALL_HEADERS = (*REQUIRED_HEADERS, OPTIONAL_CATEGORY_HEADER)

HEADER_ALIASES = {
    "主题": ("主题", "主题 Subject", "Subject"),
    "第一层": ("第一层", "第一层名称", "第一层名称 Tier1_Name", "Tier1_Name"),
    "第二层": ("第二层", "第二层名称", "第二层名称 Tier2_Name", "Tier2_Name"),
    "第三层": ("第三层", "第三层名称", "第三层名称 Tier3_Name", "Tier3_Name"),
    "第四层": ("第四层", "第四层名称", "第四层名称 Tier4_Name", "Tier4_Name"),
    "NEIC_Code": ("NEIC_Code", "国民经济行业代码", "国民经济行业代码 NEIC_Code"),
    "NEIC_Name": ("NEIC_Name", "国民经济行业名称", "国民经济行业名称 NEIC_Name"),
    OPTIONAL_CATEGORY_HEADER: (OPTIONAL_CATEGORY_HEADER,),
}


@dataclass(frozen=True)
class MappingAssetSpec:
    scenario_id: str
    filename: str
    category_name: str

    @property
    def path(self) -> Path:
        return ASSET_DIR / self.filename


@dataclass(frozen=True)
class MappingAssetReport:
    scenario_id: str
    path: Path
    source_hash: str
    data_row_count: int
    headers: tuple[str, ...]


ASSET_SPECS = (
    MappingAssetSpec("green_finance", "绿色金融.xlsx", "绿色金融"),
    MappingAssetSpec("digital_finance", "数字金融.xlsx", "数字金融"),
    MappingAssetSpec("pension_finance", "养老金融.xlsx", "养老金融"),
)


class MappingAssetValidationError(ValueError):
    """Raised when a scenario mapping workbook violates the asset contract."""


def _header_key(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip().casefold()
    return re.sub(r"\s+", "", text)


HEADER_LOOKUP = {
    _header_key(alias): canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _canonical_header(value: object) -> str | None:
    return HEADER_LOOKUP.get(_header_key(value))


def _clean_value(value: object) -> str:
    if value is None:
        return ""
    return " ".join(unicodedata.normalize("NFKC", str(value)).split())


def _is_ignored_asset(path: Path) -> bool:
    return path.name == ".DS_Store" or path.name.startswith("~$")


def discover_mapping_assets(asset_dir: Path = ASSET_DIR) -> dict[str, Path]:
    """Find the three formal workbooks while ignoring filesystem noise."""
    if not asset_dir.is_dir():
        raise MappingAssetValidationError(f"映射资产目录不存在: {asset_dir}")

    expected = {spec.filename: spec for spec in ASSET_SPECS}
    discovered: dict[str, Path] = {}
    for path in asset_dir.iterdir():
        if _is_ignored_asset(path) or not path.is_file():
            continue
        spec = expected.get(path.name)
        if spec is not None:
            discovered[spec.scenario_id] = path

    missing = [spec.filename for spec in ASSET_SPECS if spec.scenario_id not in discovered]
    if missing:
        raise MappingAssetValidationError(f"映射资产缺失: {', '.join(missing)}")
    return discovered


def validate_mapping_asset(path: Path, spec: MappingAssetSpec) -> MappingAssetReport:
    """Validate one workbook and return its reproducibility metadata."""
    if not path.is_file():
        raise MappingAssetValidationError(f"{spec.scenario_id}: 资产不存在: {path}")

    source_hash = sha256(path.read_bytes()).hexdigest()
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise MappingAssetValidationError(
            f"{spec.scenario_id}: 不是可解析的 XLSX: {path}"
        ) from exc

    try:
        rows = workbook.active.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if raw_headers is None:
            raise MappingAssetValidationError(f"{spec.scenario_id}: 工作簿没有表头")

        positions: dict[str, int] = {}
        duplicates: set[str] = set()
        for index, raw_header in enumerate(raw_headers):
            header = _canonical_header(raw_header)
            if header is None:
                continue
            if header in positions:
                duplicates.add(header)
            else:
                positions[header] = index

        missing = [header for header in REQUIRED_HEADERS if header not in positions]
        if missing or duplicates:
            details = []
            if missing:
                details.append(f"缺少表头: {', '.join(missing)}")
            if duplicates:
                details.append(f"重复表头: {', '.join(sorted(duplicates))}")
            raise MappingAssetValidationError(
                f"{spec.scenario_id}: {'; '.join(details)}"
            )

        data_row_count = 0
        category_position = positions.get(OPTIONAL_CATEGORY_HEADER)
        for source_row, row in enumerate(rows, start=2):
            if not any(
                index < len(row) and _clean_value(row[index])
                for header, index in positions.items()
                if header in REQUIRED_HEADERS
            ):
                continue
            data_row_count += 1
            if category_position is None or category_position >= len(row):
                continue
            category = _clean_value(row[category_position])
            if category and category != spec.category_name:
                raise MappingAssetValidationError(
                    f"{spec.scenario_id}: 第 {source_row} 行属于类别错配，"
                    f"应为 {spec.category_name}，实际为 {category}"
                )
    finally:
        workbook.close()

    return MappingAssetReport(
        scenario_id=spec.scenario_id,
        path=path,
        source_hash=source_hash,
        data_row_count=data_row_count,
        headers=tuple(header for header in ALL_HEADERS if header in positions),
    )


def validate_all_mapping_assets(
    asset_dir: Path = ASSET_DIR,
) -> dict[str, MappingAssetReport]:
    """Validate every formal scenario workbook in an asset directory."""
    paths = discover_mapping_assets(asset_dir)
    return {
        spec.scenario_id: validate_mapping_asset(paths[spec.scenario_id], spec)
        for spec in ASSET_SPECS
    }


def main() -> int:
    try:
        reports = validate_all_mapping_assets()
    except MappingAssetValidationError as exc:
        print(f"FAIL {exc}", file=sys.stderr)
        return 1

    for spec in ASSET_SPECS:
        report = reports[spec.scenario_id]
        print(
            f"PASS {report.scenario_id} rows={report.data_row_count} "
            f"sha256={report.source_hash}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
